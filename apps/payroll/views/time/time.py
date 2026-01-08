from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from apps.components.decorators import  role_required
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from django.contrib import messages
from apps.common.models import Tiempos ,Anos, Crearnomina , Contratos ,Nomina,Conceptosdenomina, Empresa ,Conceptosfijos ,TiemposTotales ,Sedes, Costos, Subcostos, Empresa
from django.http import HttpResponse
from django.urls import reverse
from apps.payroll.forms.SettlementForm import SettlementForm
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from datetime import datetime, date, time, timedelta
from django.db.models import Q
import pandas as pd
from django.db import transaction
from django.db.models import Sum
import holidays
from openpyxl import Workbook
from io import BytesIO
from django.db.models import F, Value, CharField
from django.db.models.functions import Concat
from apps.payroll.forms.TimeForm import TimeForm
from urllib.parse import urlencode
from openpyxl.styles import PatternFill, Font

@login_required
@role_required('accountant')
def time_list(request):
    value1 = False
    value2 = False
    value3 = False
    tiempos = []
    anio_actual = datetime.now().year
    ano_obj = Anos.objects.get(ano =  anio_actual )

    inicio_int = int(Conceptosfijos.objects.get(conceptofijo="HORARIO NOCTURNO INICIO").valorfijo)
    fin_int = int(Conceptosfijos.objects.get(conceptofijo="HORARIO NOCTURNO FIN").valorfijo)

    CO_HOLIDAYS = holidays.CO(years=ano_obj.ano)

    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']

    nominas = Crearnomina.objects.filter(
        estadonomina=True,
        id_empresa_id=idempresa
    ).order_by('-idnomina')

    selected_nomina_id = request.GET.get('datatipo')
    accion = request.GET.get('accion')
    accion2 = request.GET.get('accion2')

    contratos_empleados = (
        Contratos.objects
        .select_related('idempleado', 'idcosto', 'tipocontrato', 'idsede')
        .order_by('idempleado__papellido')
        .filter(estadocontrato=1, id_empresa=idempresa)
        .values(
            'idempleado__docidentidad', 'idempleado__papellido', 'idempleado__pnombre',
            'idempleado__snombre', 'fechainiciocontrato', 'cargo__nombrecargo', 'salario',
            'idcosto__nomcosto', 'tipocontrato__tipocontrato', 'centrotrabajo__tarifaarl',
            'idempleado__idempleado', 'idempleado__sapellido',
            'idcontrato', 'idsede__nombresede'
        )
    )
    
    empleados = [
        {
            'documento': c['idempleado__docidentidad'],
            'nombre': ' '.join(filter(None, [
                c['idempleado__papellido'],
                c['idempleado__sapellido'],
                c['idempleado__pnombre'],
                c['idempleado__snombre'],
            ])),
            'idcontrato': c['idcontrato'],
            'salario': c['salario'],
            'sede': c['idsede__nombresede'],

            # 👇 CAMPOS VACÍOS (SOLO PARA LLENAR)
            'horas_trabajadas': 0,
            'horas_normales': 0,
            'horas_domfes': 0,
            'hed': 0,
            'vhed': 0,
            'hen': 0,
            'vhen': 0,
            'hedf': 0,
            'vhedf': 0,
            'henf': 0,
            'vhenf': 0,
            'rn': 0,
            'vrn': 0,
            'rnf': 0,
            'vrnf': 0,
            'dyf': 0,
            'Vdyf':0
        }
        for c in contratos_empleados
    ]
    
    # mapa rápido por contrato
    
    empleados_map = {e['idcontrato']: e for e in empleados}
    
    # ---------- Helpers ----------
        
    # ---------- Función auxiliar para calcular extras --------

    if selected_nomina_id:
        # Traer tiempos de la nómina seleccionada
        tiempos = Tiempos.objects.filter(idnomina=selected_nomina_id).select_related(
            'idcontrato', 'idcontrato__idempleado'
        ).annotate(
            nombre_completo=Concat(
                F('idcontrato__idempleado__papellido'),
                Value(' '),
                F('idcontrato__idempleado__sapellido'),
                Value(' '),
                F('idcontrato__idempleado__pnombre'),
                Value(' '),
                F('idcontrato__idempleado__snombre'),
                output_field=CharField()
            )
        ).order_by('fechaingreso')


        for t in tiempos:

            #* =========  Variables de Inicio =========
            MINUTO_HORA = 1
            horas_trabajadas = 0.0
            horas_ordinarias = 0.0
            hed = hen = hedf = henf = rn = rnf = dyf = horas_domfes = timee = 0.0

            inicio = datetime.combine(t.fechaingreso, t.horaingreso)
            fin = datetime.combine(t.fechasalida, t.horasalida)

            actual = inicio
            paso = timedelta(minutes=1)

            noct_inicio = time(inicio_int, 0, 0)
            noct_fin = time(fin_int, 0, 0)


            inicio_diurna = noct_fin
            fin_diurna = noct_inicio

            au = actual.date()
            is_domingo = au.weekday() == 6
            is_festivo = au in CO_HOLIDAYS
            is_festivo_o_dom = is_domingo or is_festivo

            while actual < fin:
                fecha = actual.date()
                hora = actual.time()

                is_domingo = fecha.weekday() == 6
                is_festivo = fecha in CO_HOLIDAYS
                
                is_festivo_o_dom = is_domingo or is_festivo
                is_nocturna = hora < inicio_diurna or hora >= fin_diurna
                

                horas_trabajadas += MINUTO_HORA

                
                if (actual - inicio) < timedelta(minutes=480) and not is_festivo_o_dom :
                    horas_ordinarias += MINUTO_HORA

                if is_festivo_o_dom and not is_nocturna : 
                    horas_domfes += MINUTO_HORA

                if is_festivo_o_dom and is_nocturna : 
                    rnf += MINUTO_HORA
                    
                if not is_festivo_o_dom and is_nocturna : 
                    rn += MINUTO_HORA

                # Validación de 480 minutos ( 8 horas de trabajo )
                if (actual - inicio) >= timedelta(minutes=480) and not is_festivo_o_dom :
                    
                    if is_nocturna:
                        hen += MINUTO_HORA
                    else:
                        hed += MINUTO_HORA

                if (actual - inicio) >= timedelta(minutes=480) and is_festivo_o_dom :
                    
                    if is_nocturna:
                        henf += MINUTO_HORA
                    else:
                        hedf += MINUTO_HORA
                
                # avanzar
                actual += paso


            # --- Aplicar horas de descuento si existen ---
            # registro.horasdescuentos es TimeField (HH:MM:SS) -> convertir a horas float
            if t.horasdescuentos:
                h = t.horasdescuentos
                descuento_horas = h.hour + (h.minute / 60.0) + (h.second / 3600.0)
            else:
                descuento_horas = 0.0
                
            
            descuento_horas = round(descuento_horas,3)
            horas_trabajadas = round( horas_trabajadas / 60.0, 3)  

            descuento_horas = round(descuento_horas,3)
            horas_domfes = round( horas_domfes / 60.0, 3)  
            hed = round( hed / 60.0, 3)   
            hen = round( hen / 60.0, 3)               
            
            rn = round( rn / 60.0, 3)    
            rnf = round( rnf / 60.0, 3)   
            dyf = round( dyf / 60.0, 3)   
            
            horas_ordinarias = round( horas_ordinarias / 60.0, 3)  
            horas_ordinarias = max(horas_ordinarias, 0)

            if horas_trabajadas >= descuento_horas:
                horas_trabajadas -= descuento_horas

            if horas_domfes >= descuento_horas:
                horas_domfes -= descuento_horas

            if horas_ordinarias >= descuento_horas:
                horas_ordinarias -= descuento_horas


            h_totales = horas_trabajadas + horas_domfes
            print(f" hd {horas_domfes} ht {horas_trabajadas} ho {horas_ordinarias}")


            if h_totales >= 8:
                hed = hed
            else :
                hed = 0

            if h_totales >= 8:
                hen = hen
            else :
                hen = 0


            if hed >= descuento_horas:
                hed -= descuento_horas

            if hen >= descuento_horas:
                hen -= descuento_horas


            if hedf >= descuento_horas:
                hedf -= descuento_horas

            if henf >= descuento_horas:
                henf -= descuento_horas

            if horas_domfes > descuento_horas:
                horas_domfes -= (descuento_horas) 
                dyf = horas_domfes

            if rn > 8:
                rn -= (descuento_horas) 

            if rnf > 8:
                rnf -= (descuento_horas)  

            t.horas_trabajadas= round( horas_trabajadas, 3) 
            t.horas_ordinarias= round( horas_ordinarias, 3)  
            t.saldo_horas= 0
            t.horas_dom_fest= round( horas_domfes, 3)  

            t.hed = round( hed, 3)   
            t.hen = round( hen, 3)   
            t.hedf = round( hedf, 3)  
            t.henf = round( henf, 3)  
            t.rn = round( rn, 3)  
            t.rnf = round( rnf, 3)  
            t.dyf = round( dyf, 3)   

        value1 = True

    ##! ===================== Calcular =====================
    
    if accion == 'calcular' and selected_nomina_id:

        value2 = True
        value1 = False
        

        tiempos = Tiempos.objects.filter(
            idnomina=selected_nomina_id
        ).select_related('idcontrato__idsede')
        

        conceptos = Conceptosfijos.objects.filter(
            conceptofijo__in=[
                'JORNADA MAXIMA MENSUAL',
                'HORA EXTRA DIURNA FACTOR',
                'HORA EXTRA NOCTURNA FACTOR',
                'HORA EXTRA DIURNA FESTIVA FACTOR',
                'HORA EXTRA NOCTURNA FESTIVA FACTOR',
                'RECARGO NOCTURNO FACTOR',
                'RECARGO DOMINICAL O FESTIVO FACTOR',
                
            ]
        ).values('conceptofijo', 'valorfijo')

        factores = {c['conceptofijo']: float(c['valorfijo']) for c in conceptos}

        jmm = factores.get('JORNADA MAXIMA MENSUAL', 220.0)  # ejemplo 220h/mes
        hed_factor  = factores.get('HORA EXTRA DIURNA FACTOR', 1.0)
        hen_factor  = factores.get('HORA EXTRA NOCTURNA FACTOR', 1.0)
        hedf_factor = factores.get('HORA EXTRA DIURNA FESTIVA FACTOR', 1.0)
        henf_factor = factores.get('HORA EXTRA NOCTURNA FESTIVA FACTOR', 1.0)
        rn_factor   = factores.get('RECARGO NOCTURNO FACTOR', 1.0)
        rdf_factor  = factores.get('RECARGO DOMINICAL O FESTIVO FACTOR', 1.0)
        
        
        for registro in tiempos.order_by('idcontrato__idcontrato'):

            contrato_id = registro.idcontrato.idcontrato
            emp = empleados_map.get(contrato_id)

            if not emp:
                continue

            #* =========  Variables de Inicio =========
            MINUTO_HORA = 1
            horas_trabajadas = 0.0
            horas_ordinarias = 0.0
            hed = hen = hedf = henf = rn = rnf = dyf = horas_domfes = timee = 0.0

            inicio = datetime.combine(registro.fechaingreso, registro.horaingreso)
            fin = datetime.combine(registro.fechasalida, registro.horasalida)

            actual = inicio
            paso = timedelta(minutes=1)

            noct_inicio = time(inicio_int, 0, 0)
            noct_fin = time(fin_int, 0, 0)


            inicio_diurna = noct_fin
            fin_diurna = noct_inicio


            
            au = actual.date()
            is_domingo = au.weekday() == 6
            is_festivo = au in CO_HOLIDAYS
            is_festivo_o_dom = is_domingo or is_festivo

            while actual < fin:
                fecha = actual.date()
                hora = actual.time()

                is_domingo = fecha.weekday() == 6
                is_festivo = fecha in CO_HOLIDAYS
                
                is_festivo_o_dom = is_domingo or is_festivo
                is_nocturna = hora < inicio_diurna or hora >= fin_diurna
                

                horas_trabajadas += MINUTO_HORA

                
                if (actual - inicio) < timedelta(minutes=480) and not is_festivo_o_dom :
                    horas_ordinarias += MINUTO_HORA

                if is_festivo_o_dom and not is_nocturna : 
                    horas_domfes += MINUTO_HORA

                if is_festivo_o_dom and is_nocturna : 
                    rnf += MINUTO_HORA
                    
                if not is_festivo_o_dom and is_nocturna : 
                    rn += MINUTO_HORA

                # Validación de 480 minutos ( 8 horas de trabajo )
                if (actual - inicio) >= timedelta(minutes=480) and not is_festivo_o_dom :
                    
                    if is_nocturna:
                        hen += MINUTO_HORA
                    else:
                        hed += MINUTO_HORA

                if (actual - inicio) >= timedelta(minutes=480) and is_festivo_o_dom :
                    
                    if is_nocturna:
                        henf += MINUTO_HORA
                    else:
                        hedf += MINUTO_HORA
                
                # avanzar
                actual += paso


            # --- Aplicar horas de descuento si existen ---
            # registro.horasdescuentos es TimeField (HH:MM:SS) -> convertir a horas float
            if registro.horasdescuentos:
                h = registro.horasdescuentos
                descuento_horas = h.hour + (h.minute / 60.0) + (h.second / 3600.0)
            else:
                descuento_horas = 0.0
                
            
            descuento_horas = round(descuento_horas,3)
            horas_trabajadas = round( horas_trabajadas / 60.0, 3)  

            descuento_horas = round(descuento_horas,3)
            horas_domfes = round( horas_domfes / 60.0, 3)  
            hed = round( hed / 60.0, 3)   
            hen = round( hen / 60.0, 3)               
            
            rn = round( rn / 60.0, 3)    
            rnf = round( rnf / 60.0, 3)   
            dyf = round( dyf / 60.0, 3)   
            
            horas_ordinarias = round( horas_ordinarias / 60.0, 3)  
            horas_ordinarias = max(horas_ordinarias, 0)

            if horas_trabajadas >= descuento_horas:
                horas_trabajadas -= descuento_horas

            if horas_domfes >= descuento_horas:
                horas_domfes -= descuento_horas

            if horas_ordinarias >= descuento_horas:
                horas_ordinarias -= descuento_horas


            h_totales = horas_trabajadas + horas_domfes


            if h_totales >= 8:
                hed = hed
            else :
                hed = 0

            if h_totales >= 8:
                hen = hen
            else :
                hen = 0


            if hed >= descuento_horas:
                hed -= descuento_horas

            if hen >= descuento_horas:
                hen -= descuento_horas


            if hedf >= descuento_horas:
                hedf -= descuento_horas

            if henf >= descuento_horas:
                henf -= descuento_horas

            if horas_domfes > descuento_horas:
                horas_domfes -= (descuento_horas) 
                dyf = horas_domfes

            if rn > 8:
                rn -= (descuento_horas) 

            if rnf > 8:
                rnf -= (descuento_horas)  


            # =============================
            # ✅ SOLO APPEND A EMPLEADOS
            # =============================
            
            #print(f" fecha {inicio} - hed : {hed}  " )
            
            emp['horas_trabajadas'] += horas_trabajadas
            emp['horas_normales'] += horas_ordinarias
            emp['horas_domfes'] += horas_domfes
            emp['hed'] += hed
            emp['hen'] += hen
            emp['hedf'] += hedf
            emp['henf'] += henf
            emp['rn'] += rn
            emp['rnf'] += rnf
            emp['dyf'] += horas_domfes
            

        for e in empleados:
            aux = ( e['salario'] / jmm )
            e['horas_trabajadas'] = round(e['horas_trabajadas'], 3)
            e['horas_normales'] = round(e['horas_normales'], 3)
            e['horas_domfes'] = round(e['horas_domfes'], 3)
            e['dyf'] = round(e['horas_domfes'], 3)

            e['hed'] = round(e['hed'], 3)
            e['Vhed'] = e['hed'] * aux * hed_factor 
            
            e['hen'] = round(e['hen'], 3)
            e['Vhen'] = e['hen'] * aux * hen_factor 
            
            e['hedf'] = round(e['hedf'], 3)
            e['Vhedf'] = e['hedf'] * aux * hedf_factor 
            
            e['henf'] = round(e['henf'], 3)
            e['vhenf'] = e['henf'] * aux * henf_factor 
            
            e['rn'] = round(e['rn'], 3)
            e['Vrn'] = e['rn'] * aux * rn_factor 
            
            e['rnf'] = round(e['rnf'], 3)
            e['Vrnf'] = e['rnf'] * aux * rdf_factor 
            
            
            e['dyf'] = round(e['rnf'], 3)
            e['Vdyf'] = e['dyf'] * aux * rdf_factor 
            
    
    ##! ===================== Grabar =====================
    
    if accion2 == 'grabar':
        
        value1 = False
        value2 = True
        value3 = True
        
        for data in empleados:
            contrato = Contratos.objects.get(idcontrato=data['idcontrato'])
            nomina = Crearnomina.objects.get(pk=selected_nomina_id)
            costo = Costos.objects.get(pk=contrato.idcosto.idcosto)
            empresa = Empresa.objects.get(pk=idempresa)

            horas_ord = Decimal(data.get('horas_normales') or 0)
            horas_trab = Decimal(data.get('horas_trabajadas') or 0)

            # 🔴 Si ambas son 0 → no crear ni actualizar
            if horas_ord == 0 and horas_trab == 0:
                continue  # o `continue` si estás dentro de un loop
            
            # 🔹 Buscar o crear el registro
            obj, created = TiemposTotales.objects.update_or_create(
                idcontrato=contrato,
                idnomina=nomina,
                defaults={
                    'horasord':data.get('horas_normales', 0),
                    'horastrab':data.get('horas_trabajadas', 0),
                    'hed':data.get('hed', 0),
                    'vhed':data.get('Vhed', 0),
                    'hen':data.get('hen', 0),
                    'vhen':data.get('Vhen', 0),
                    'hedf':data.get('hedf', 0),
                    'vhedf':data.get('Vhedf', 0),
                    'henf':data.get('henf', 0),
                    'vhenf': data.get('Vhenf', 0),
                    'rn': data.get('rn', 0),
                    'vrn': data.get('Vrn', 0),
                    'rnf': data.get('rnf', 0),
                    'vrnf': data.get('Vrnf', 0),
                    'dyf': data.get('dyf', 0),  
                    'vdyf': data.get('Vdyf', 0),
                    'valorextras': data.get('ValorExtras', 0),
                    'idcosto': costo,
                    'idempresa': empresa,
                }
            )
            
            

        messages.success(request, 'Tiempos guardados correctamente')

        
    if accion2 == 'grabar2':
        value1 = value2 = value3 = False
        nomina = Crearnomina.objects.get(pk=selected_nomina_id)

        # 🔹 Cargar conceptos masivamente
        codigos = [3, 5, 6, 7, 8, 9, 16]
        conceptos = {
            c.codigo: c for c in Conceptosdenomina.objects.filter(codigo__in=codigos , id_empresa=idempresa )
        }

        
        faltantes = [c for c in codigos if c not in conceptos]
        if faltantes:
            messages.error(request, f"Faltan conceptos en la empresa: {faltantes}")
            return redirect(request.path)

        # 🔹 Relación campo ↔ concepto
        campos_a_concepto = {
            'hed': conceptos[5],
            'hen': conceptos[6],
            'dyf': conceptos[7],
            'hedf': conceptos[8],
            'henf': conceptos[9],
            'rn': conceptos[3],
            'rnf': conceptos[16],
        }

        tiempos = TiemposTotales.objects.filter(idnomina_id=selected_nomina_id)



        
        for data in tiempos:
            for campo, concepto in campos_a_concepto.items():
                cantidad = getattr(data, campo, 0)
                valor_campo = f"v{campo}" if hasattr(data, f"v{campo}") else None
                valor = getattr(data, valor_campo, 0) if valor_campo else 0
                
                try:
                    # 🔹 Normalizar valores None
                    cantidad = Decimal(cantidad) if cantidad is not None else Decimal('0')

                    # 🔹 Redondear a 2 decimales
                    cantidad = cantidad.quantize(Decimal('0.01'))
                    valor = Decimal(data.idcontrato.salario / jmm ) * Decimal(cantidad) * Decimal(concepto.multiplicadorconcepto)

                    # 🔹 Validar rango
                    if cantidad > 0 and valor < Decimal('9999999999.99'):
                        obj, created = Nomina.objects.update_or_create(
                            idcontrato=data.idcontrato,
                            idnomina=nomina,
                            idconcepto=concepto,
                            defaults={
                                'cantidad': cantidad,
                                'valor': valor,
                                'estadonomina': 1,
                            },
                        )
                        


                except (InvalidOperation, OverflowError, TypeError) as e:
                    print(f"❌ Error numérico en contrato {data.idcontrato}: {e}")

        messages.success(request, "Tiempos aplicados correctamente a la nómina.")
        return redirect(reverse('payroll:payrollview', args=[selected_nomina_id]))

            
            
        
    return render(
        request,
        './payroll/time_list.html',
        {
            'tiempos': tiempos,
            'nominas': nominas,
            'true1': value1,
            'true2': value2,
            'true3': value3,
            'empleados': empleados,
            'selected_nomina_id': selected_nomina_id
        }
    )





@login_required
@role_required('accountant')
def time_edit(request,id):
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']
    
    tiempo = Tiempos.objects.get(idmarcacion = id)
    
    initial={
        'fechaingreso': tiempo.fechaingreso,
        'fechasalida': tiempo.fechasalida,
        'horaingreso': tiempo.horaingreso,
        'horasalida': tiempo.horasalida,
        'horasdescuentos': tiempo.horasdescuentos.hour if tiempo.horasdescuentos else 0,
        'contract': tiempo.idcontrato.idcontrato
    }
    
    form = TimeForm(idempresa = idempresa , initial = initial , id = id )
    
    if request.method == 'POST':
        form = TimeForm(request.POST , idempresa=idempresa , id = id)
        if form.is_valid():
            
            fechaingreso = form.cleaned_data['fechaingreso']
            fechasalida = form.cleaned_data['fechasalida']
            horaingreso = form.cleaned_data['horaingreso']
            horasalida = form.cleaned_data['horasalida']
            horasdescuentos = form.cleaned_data['horasdescuentos']
            
            if tiempo.fechaingreso != fechaingreso:
                tiempo.fechaingreso = fechaingreso 
            
            if tiempo.fechasalida != fechasalida:
                tiempo.fechasalida = fechasalida 
                
            if tiempo.horaingreso != horaingreso:
                tiempo.horaingreso = horaingreso 
                
            if tiempo.horasalida != horasalida:
                tiempo.horasalida = horasalida 
                
            if horasdescuentos is not None:
                nueva_hora = time(hour=int(horasdescuentos), minute=0)
                if tiempo.horasdescuentos != nueva_hora:
                    tiempo.horasdescuentos = nueva_hora

            
            tiempo.save()
            
            response = HttpResponse()
            response['X-Up-Accept-Layer'] = 'true'  #Indica a Unpoly que acepte (cierre) el modal
            response['X-Up-icon'] = 'success'  # URL para recargar la página principal   
            response['X-Up-message'] = 'Tiempo actualizado exitosamente'    
            
            params = {'datatipo': tiempo.idnomina.idnomina }
            url_base = reverse('payroll:time_list')  # genera algo como "/payroll/time/"
            url_completa = f"{url_base}?{urlencode(params)}"
            response['X-Up-Location'] = url_completa
            
            return response
        
    
    return render(
        request,
        './payroll/partials/time_edit.html',
        {'form': form}
    )


@login_required
@role_required('accountant')
def time_doc(request, id):
    usuario = request.session.get('usuario', {})
    idempresa = usuario.get('idempresa')

    tiempos = Tiempos.objects.filter(idnomina=id).select_related('idcontrato__idsede')
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Tiempo Marcados"

    headers = [
        'Idcontrato', 'Empleado', 'FechaIngreso', 'HoraIngreso', 'FechaSalida', 'HoraSalida',
        'HorasDescuentos', 'HorasTraba', 'HorasOrdi', 'Saldo', 'HorasDomFes',
        'Hed', 'Hen', 'Hedf', 'Henf', 'Rn', 'Rnf', 'Dyf', 'Sede'
    ]
    ws.append(headers)

    colors = ["FFF9C4", "C8E6C9", "BBDEFB", "FFCDD2", "D1C4E9"]
    total_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    total_font = Font(bold=True)

    current_contract = None
    color_index = 0

    CO_HOLIDAYS = holidays.CO(years=2025)

    # Inicializar acumuladores
    acumulados = {
        'HorasTraba': 0,
        'HorasOrdi': 0,
        'HorasDomFes': 0,
        'Hed': 0,
        'Hen': 0,
        'Hedf': 0,
        'Henf': 0,
        'Rn': 0,
        'Rnf': 0,
        'Dyf': 0
    }

    
    
    

    def agregar_totales_contrato(contrato_id):
        """Agrega una fila con los totales acumulados del contrato actual."""
        if contrato_id is None:
            return
        total_row = [
            contrato_id,
            "TOTAL CONTRATO",
            "", "", "", "",
            "",
            acumulados['HorasTraba'],
            acumulados['HorasOrdi'],
            "",
            acumulados['HorasDomFes'],
            acumulados['Hed'], acumulados['Hen'], acumulados['Hedf'], acumulados['Henf'],
            acumulados['Rn'], acumulados['Rnf'], acumulados['Dyf'],
            "",
        ]
        ws.append(total_row)

        # Estilos
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font

        # Reiniciar acumuladores
        for k in acumulados.keys():
            acumulados[k] = 0

    # 🔁 Recorrer registros
    for registro in tiempos.order_by('idcontrato__idcontrato'):
        contrato_id = registro.idcontrato.idcontrato

        # Si cambia el contrato, agregamos totales del anterior
        if current_contract is not None and contrato_id != current_contract:
            agregar_totales_contrato(current_contract)

        # Cambiar color del bloque
        if contrato_id != current_contract:
            current_contract = contrato_id
            color_index = (color_index + 1) % len(colors)
            fill = PatternFill(start_color=colors[color_index],
                                end_color=colors[color_index],
                                fill_type="solid")
        MINUTO_HORA = 1
        
        # Inicializar contadores (en horas)
        horas_trabajadas = 0.0
        horas_ordinarias = 0.0
        hed = hen = hedf = henf = rn = rnf = dyf = horas_domfes = timee = 0.0

        inicio = datetime.combine(registro.fechaingreso, registro.horaingreso)
        fin = datetime.combine(registro.fechasalida, registro.horasalida)

        actual = inicio
        paso = timedelta(minutes=1)

        # Rangos diurnos (ajusta si tu política es otra)
        
        inicio_int = int(Conceptosfijos.objects.get(conceptofijo="HORARIO NOCTURNO INICIO").valorfijo)
        fin_int = int(Conceptosfijos.objects.get(conceptofijo="HORARIO NOCTURNO FIN").valorfijo)

        noct_inicio = time(inicio_int, 0, 0)
        noct_fin = time(fin_int, 0, 0)
        # Diurna es el complemento del rango nocturno
        inicio_diurna = noct_fin
        fin_diurna = noct_inicio

        au = actual.date()
        is_domingo = au.weekday() == 6
        is_festivo = au in CO_HOLIDAYS
        is_festivo_o_dom = is_domingo or is_festivo
        
        
        while actual < fin:
            fecha = actual.date()
            hora = actual.time()

            is_domingo = fecha.weekday() == 6
            is_festivo = fecha in CO_HOLIDAYS
            
            is_festivo_o_dom = is_domingo or is_festivo
            is_nocturna = hora < inicio_diurna or hora >= fin_diurna
            

            horas_trabajadas += MINUTO_HORA

            
            if (actual - inicio) < timedelta(minutes=480) and not is_festivo_o_dom :
                horas_ordinarias += MINUTO_HORA

            if is_festivo_o_dom and not is_nocturna : 
                horas_domfes += MINUTO_HORA

            if is_festivo_o_dom and is_nocturna : 
                rnf += MINUTO_HORA
                
            if not is_festivo_o_dom and is_nocturna : 
                rn += MINUTO_HORA

            # Validación de 480 minutos ( 8 horas de trabajo )
            if (actual - inicio) >= timedelta(minutes=480) and not is_festivo_o_dom :
                
                if is_nocturna:
                    hen += MINUTO_HORA
                else:
                    hed += MINUTO_HORA

            if (actual - inicio) >= timedelta(minutes=480) and is_festivo_o_dom :
                
                if is_nocturna:
                    henf += MINUTO_HORA
                else:
                    hedf += MINUTO_HORA
            
            # avanzar
            actual += paso


        # --- Aplicar horas de descuento si existen ---
        # registro.horasdescuentos es TimeField (HH:MM:SS) -> convertir a horas float
        if registro.horasdescuentos:
            h = registro.horasdescuentos
            descuento_horas = h.hour + (h.minute / 60.0) + (h.second / 3600.0)
        else:
            descuento_horas = 0.0
            
        
        descuento_horas = round(descuento_horas,3)
        horas_trabajadas = round( horas_trabajadas / 60.0, 3)  

        descuento_horas = round(descuento_horas,3)
        horas_domfes = round( horas_domfes / 60.0, 3)  
        hed = round( hed / 60.0, 3)   
        hen = round( hen / 60.0, 3)               
        
        rn = round( rn / 60.0, 3)    
        rnf = round( rnf / 60.0, 3)   
        dyf = round( dyf / 60.0, 3)   
        
        horas_ordinarias = round( horas_ordinarias / 60.0, 3)  
        horas_ordinarias = max(horas_ordinarias, 0)

        if horas_trabajadas >= descuento_horas:
            horas_trabajadas -= descuento_horas

        if horas_domfes >= descuento_horas:
            horas_domfes -= descuento_horas

        if horas_ordinarias >= descuento_horas:
            horas_ordinarias -= descuento_horas


        h_totales = horas_trabajadas + horas_domfes


        if h_totales >= 8:
            hed = hed
        else :
            hed = 0

        if h_totales >= 8:
            hen = hen
        else :
            hen = 0


        if hed >= descuento_horas:
            hed -= descuento_horas

        if hen >= descuento_horas:
            hen -= descuento_horas


        if hedf >= descuento_horas:
            hedf -= descuento_horas

        if henf >= descuento_horas:
            henf -= descuento_horas

        if horas_domfes > descuento_horas:
            horas_domfes -= (descuento_horas) 
            dyf = horas_domfes

        if rn > 8:
            rn -= (descuento_horas) 

        if rnf > 8:
            rnf -= (descuento_horas)  

        # Acumular totales
        acumulados['HorasTraba'] += horas_trabajadas 
        acumulados['HorasOrdi'] += horas_ordinarias
        acumulados['HorasDomFes'] += horas_domfes
        acumulados['Hed'] += hed
        acumulados['Hen'] += hen
        acumulados['Hedf'] += hedf
        acumulados['Henf'] += henf
        acumulados['Rn'] += rn
        acumulados['Rnf'] += rnf
        acumulados['Dyf'] += dyf

        # Agregar fila
        row = [
            contrato_id,
            getattr(registro.idcontrato, 'empleado', 'N/A'),
            registro.fechaingreso,
            registro.horaingreso,
            registro.fechasalida,
            registro.horasalida,
            registro.horasdescuentos,
            horas_trabajadas,
            horas_ordinarias,
            '0.0',
            horas_domfes,
            hed, hen, hedf, henf, rn, rnf, dyf,
            registro.idcontrato.idsede.nombresede if registro.idcontrato.idsede else "",
        ]

        
        ws.append(row)

        for cell in ws[ws.max_row]:
            cell.fill = fill

        ws.cell(row=ws.max_row, column=3).number_format = 'DD/MM/YYYY'
        ws.cell(row=ws.max_row, column=5).number_format = 'DD/MM/YYYY'

    # 🔚 Totales del último contrato
    agregar_totales_contrato(current_contract)

    output = BytesIO()
    wb.save(output)
    wb.close() 
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Tiempo_Marcados.xlsx"'
    
    # Evitar caché
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    
    return response
    


def formatear_fecha(valor):
    if isinstance(valor, datetime):   # Si ya es datetime
        return valor.date()
    if isinstance(valor, str) and '/' in valor:
        try:
            return datetime.strptime(valor, '%d/%m/%Y').date()
        except:
            return valor
    return valor



def time_add(request):
    usuario = request.session.get('usuario', {})
    idempresa = usuario.get('idempresa')

    nominas = Crearnomina.objects.filter(
        estadonomina=True,
        id_empresa_id=idempresa
    ).order_by('-idnomina')

    if request.method == 'POST' and request.FILES.get('file'):
        errors = []
        file = request.FILES['file']
        idnomina = request.POST.get('idnomina')

        try:
            df = pd.read_csv(file, header=None,sep=";", encoding="utf-8")
        except Exception as e:
            errors.append(f"Error al leer el archivo: {str(e)}")
            return render(request, './companies/partials/disability_upload_errors.html', {'errors': errors})
        df = df.dropna(how="all")
        
        
        registros_validados = []

        for idx, fila in df.iterrows():
            try:
                contrato      = fila[0]
                fecha_ingreso = formatear_fecha(fila[1])
                fecha_salida  = formatear_fecha(fila[2])
                hora_ingreso  = fila[3]
                hora_salida   = fila[4]
                horas_descuentos  = fila[5]

                # --- Validaciones ---
                if not contrato:
                    errors.append(f"Fila {idx+1}: El contrato es obligatorio.")
                    continue

                if not fecha_ingreso:
                    errors.append(f"Fila {idx+1}: La fecha de ingreso no es válida.")
                    continue

                if fecha_salida and fecha_salida < fecha_ingreso:
                    errors.append(f"Fila {idx+1}: La fecha de salida no puede ser menor que la de ingreso.")
                    continue

                if not hora_ingreso or not hora_salida:
                    errors.append(f"Fila {idx+1}: Hora de ingreso y salida son obligatorias.")
                    continue

                empresa = Empresa.objects.get(idempresa =  idempresa )
                
                # Verificar que el contrato exista
                if not Contratos.objects.filter(old_idcontrato=contrato, id_empresa=idempresa).exists():
                    errors.append(f"Fila {idx+1}: El contrato {contrato} no existe en la empresa {empresa.nombreempresa}.")
                    continue

                # Evitar duplicados: contrato + fecha ingreso ya existente
                # if Tiempos.objects.filter(idcontrato__old_idcontrato=contrato, fechaingreso=fecha_ingreso, idempresa_id=idempresa).exists():
                #     errors.append(f"Fila {idx+1}: Ya existe un registro de tiempo para contrato {contrato} en la fecha {fecha_ingreso}.")
                #     continue
                
                
                
                
                contr = Contratos.objects.filter(old_idcontrato=contrato, id_empresa=idempresa).first()
                
                
                
                # Si pasa todas las validaciones, lo agregamos a lista
                tiempo, created = Tiempos.objects.update_or_create(
                    # 🔍 Campos para buscar si ya existe
                    fechaingreso=fecha_ingreso,
                    idcontrato=contr,
                    idnomina_id=idnomina,
                    idempresa_id=idempresa,

                    # ✏️ Campos que se actualizan si existe
                    defaults={
                        'horaingreso': hora_ingreso,
                        'fechasalida': fecha_salida,
                        'horasalida': hora_salida,
                        'horasdescuentos': horas_descuentos,
                    }
                )

                
                print(created)
            except Exception as e:
                errors.append(f"Fila {idx+1}: Error inesperado -> {str(e)}")

        # Si hubo errores: no se guarda nada
        if errors:
            return render(request, './companies/partials/disability_upload_errors.html', {
                'errors': errors
            })
        else:
            return render(request, 'payroll/partials/success_time.html')
        

    return render(request, 'payroll/partials/time_add.html', {'nominas': nominas})




def time_data(request,id):
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']
    
    times = Tiempos.objects.filter(idempresa = idempresa ,idcontrato = id ).order_by('-fechaingreso') 
    
    return render(request, './payroll/partials/time_data.html', {'times': times})
