from django.shortcuts import render
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required
from apps.components.decorators import role_required
import pandas as pd
from openpyxl.utils.exceptions import InvalidFileException
from apps.common.models import Conceptosdenomina, Contratos
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.http import JsonResponse

# Diccionario de mensajes de error con números como claves
error_messages = {
    '1': "Error code 1001: Id contrato inválido o faltante",
    '2': "Error code 1002: Concepto faltante o no es un número entero",
    '3': "Error code 1003: Contrato no encontrado",
    '4': "Error code 1004: El estado del contrato no es válido (debe estar activo)",
    '5': "Error code 1005: Concepto no encontrado",
    '6': "Error code 1006: El contrato no pertenece a la empresa actual",
    '7': "Error code 1007: Fila vacía o incompleta",
    '8': "Error code 1008: El valor debe ser mayor que cero",
    '9': "Error code 1009: Concepto duplicado para este contrato",
    '10': "Error code 1010: Faltan valores críticos (Id contrato, Concepto)",
    '11': "Error code 1012: La cantidad no puede ser negativa",
    '12': "Error code 1013: El contrato está cerrado o inactivo",
    '13': "Error code 1014: El valor no puede ser negativo",
    '15': "Error code 1015: El contrato y concepto ya existen",
    'general': "Error general: Ocurrió un error procesando el archivo",
    "16": "Los siguientes conceptos no están reconocidos: {concepts}",
    "17": "No se proporcionaron conceptos para evaluar.",
    "18": "Tipo de entrada no válido. Se esperaba una lista de conceptos.",
}

def validate_concepts(df, available_concepts):
    errors = []  # Lista para acumular errores

    # Intentar convertir los índices de las columnas a enteros, excepto 'Contrato' y 'Nombre'
    for col in df.columns:
        if col not in ['Contrato', 'Nombre']:
            try:
                # Verificar si todos los valores de la columna son numéricos
                df[col] = pd.to_numeric(df[col], errors='raise')  # Forzar conversión a números
            except ValueError as e:
                # Si no se puede convertir, agregamos el error a la lista
                errors.append(f"Error: La columna '{col}' no se puede convertir a entero. Detalles: {e}")

    # Validar que available_concepts sea una lista de enteros
    if not isinstance(available_concepts, list) or not all(isinstance(c, int) for c in available_concepts):
        errors.append(error_messages["18"])

    if not available_concepts:
        errors.append(error_messages["17"])

    # Convertir nombres de columnas a enteros si es posible
    columns_df = []
    for col in df.columns:
        if col not in ['Contrato', 'Nombre']:
            try:
                columns_df.append(int(col))  # Intentar convertir cada nombre de columna a entero
            except ValueError:
                # Si no se puede convertir, agregamos el error a la lista
                col = col.replace('Unnamed: ', '')  # Eliminar 'Unnamed: ' si es una columna sin nombre
                errors.append(f"""
                        <div class="validation-error" style="
                            border-left: 4px solid #e74c3c;
                            background: #fef6f6;
                            padding: 15px;
                            margin: 10px 0;
                            border-radius: 0 4px 4px 0;
                            font-family: Arial, sans-serif;
                            color: #333;
                        ">
                            <div style="
                                display: flex;
                                align-items: center;
                                margin-bottom: 10px;
                                color: #e74c3c;
                                font-weight: bold;
                            ">
                                <svg style="margin-right: 8px; width: 18px; height: 18px;" fill="none" stroke="currentColor" viewBox="0 0 24 24" xmlns="http://www.w3.org/2000/svg">
                                    <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M12 8v4m0 4h.01M21 12a9 9 0 11-18 0 9 9 0 0118 0z"></path>
                                </svg>
                                Error de validación en Excel
                            </div>
                            
                            <p style="margin: 8px 0;">La columna <strong>{col}</strong> no cumple con el formato esperado:</p>
                            
                            <ul style="
                                margin: 8px 0;
                                padding-left: 20px;
                            ">
                                <li style="margin-bottom: 5px;">Todos los valores deben ser <strong>números enteros</strong> (sin decimales o texto)</li>
                                <li style="margin-bottom: 5px;">No debe contener celdas vacías o datos incorrectos</li>
                            </ul>
                            
                            <div style="
                                background: #fff;
                                border-radius: 4px;
                                padding: 10px;
                                margin: 10px 0;
                                border: 1px solid #eee;
                            ">
                                <div style="font-weight: bold; margin-bottom: 5px;">Revisar:</div>
                                <ul style="
                                    margin: 0;
                                    padding-left: 20px;
                                    color: #555;
                                ">
                                    <li>Formato de celdas (deben ser numéricas)</li>
                                    <li>Datos ingresados (evitar texto o caracteres no válidos)</li>
                                </ul>
                            </div>
                            
                            <div style="
                                margin-top: 12px;
                                font-size: 14px;
                            ">
                                <span style="font-weight: bold;">Solución:</span> Corrija los valores en la columna <strong>{col}</strong> y vuelva a validar.
                            </div>
                        </div>
                        """)

    # Encontrar conceptos faltantes
    missing = [col for col in columns_df if col not in available_concepts]

    if missing:
        errors.append(error_messages["16"].format(concepts=", ".join(map(str, missing))))

    # Si hubo errores, retornamos todos
    if errors:
        return "\n".join(errors)
    
    return "Todos los conceptos son válidos."



@login_required
@role_required('accountant')
def flat(request, id):
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']
    
    # Diccionario para almacenar los errores por fila
    errors = [] 
    general_error = []
    
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        if not file.name.endswith('.xlsx'):
            errors.append({'general': 'Solo se aceptan archivos con extensión .xlsx.'})
            return render(request, './payroll/plane.html', {'id': id, 'errors': errors})
        
        try:
            # Leer el archivo directamente sin guardarlo
            df = pd.read_excel(file)

            # Definir las columnas fijas
            fixed_columns = ['Contrato', 'Nombre']

            # Identificar las columnas dinámicas
            dynamic_columns = [col for col in df.columns if col not in fixed_columns]
    
            # Validar que las columnas fijas existan
            missing_columns = [col for col in fixed_columns if col not in df.columns]
            
            if missing_columns:
                errors.append({'general': f'Faltan las siguientes columnas obligatorias: {", ".join(missing_columns)}.'})
                return render(request, './payroll/plane.html', {'id': id, 'errors': errors})

            conceptos = Conceptosdenomina.objects.filter(id_empresa_id=idempresa).values_list('codigo', flat=True)

            general_error.append(validate_concepts(df,list(conceptos)))
            
            
            # Validar filas y generar errores
            for index, row in df.iterrows():
                row_errors = [] 

                # Validar las columnas fijas
                idcontrato = row.get('Contrato')
                
                if pd.isnull(idcontrato):
                    row_errors.append('1')
                else : 
                    contrato = Contratos.objects.filter(idcontrato=idcontrato, id_empresa_id=idempresa).first()                
                
                if contrato is None:
                    row_errors.append('3')
                else:
                    if contrato.estadocontrato != '1':
                        row_errors.append('4')
                        
                        
                    if contrato.id_empresa_id != idempresa:
                        row_errors.append('6')

                

                
                # for col in dynamic_columns:
                    
                #     concepto = Conceptosdenomina.objects.filter(codigo=col, id_empresa_id=idempresa).values('codigo').first()

                #     # Acceder al valor de 'codigo' directamente
                #     if concepto:
                #         codigo = concepto.get('codigo')
                #         print(codigo)
                        
                        
                #     if concepto is None:
                #         row_errors.append('5')
                
                # Agregar errores por fila si existen
                if row_errors:
                    errors.append({
                        'line': index + 1,  # Fila 1 en Excel corresponde al índice 0 en DataFrame
                        'contract_id': contrato.idcontrato  if contrato is not None else "Contrato no disponible" ,
                        'identification': contrato.idempleado.docidentidad if contrato is not None else "Identificación no disponible",
                        'name':  f"{contrato.idempleado.pnombre} {contrato.idempleado.papellido}" if contrato is not None else "Nombre no disponible" ,  # Suponiendo que no se tiene la columna Nombre
                        'error': "".join([f"<li>{error_messages[err]}</li>" for err in row_errors]) , 
                        'details' : 'detalles' 
                    })
                
                
        except Exception as e:
            errors.append({'general': f'Ocurrió un error procesando el archivo: {str(e)}'})
        
    formatted_errors = []
    for error in general_error:
        if isinstance(error, str):
            # Dividir por saltos de línea si el error es un string con saltos de línea
            formatted_errors.extend(error.split("\n"))
        else:
            formatted_errors.append(error)
            
            
    return render(request, './payroll/plane.html', {
        'id': id,
        'errors': errors,
        'general_error': formatted_errors , 
    })





def document(request):
    # Crear el libro de Excel
    wb = Workbook()

    # Hoja principal: Datos
    ws = wb.active
    ws.title = "Datos"

    # Definir las columnas fijas
    fixed_columns = ["Contrato", "Nombre"]
    # Definir las columnas dinámicas
    column_names = ["15", "10", "11"]

    # Escribir las columnas fijas
    for col_num, col_name in enumerate(fixed_columns, 1):
        ws.cell(row=1, column=col_num, value=col_name)

    # Escribir las columnas dinámicas
    for col_num, col_name in enumerate(column_names, len(fixed_columns) + 1):
        # Convertir el valor a entero si no es una columna fija
        if col_name not in fixed_columns:
            col_name = int(col_name)  # Asegúrate de que se convierte en entero
        ws.cell(row=1, column=col_num, value=col_name)

    # Aplicar estilo para que los encabezados estén centrados
    for col in range(1, len(fixed_columns) + len(column_names) + 1):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Fijar las primeras dos columnas
    ws.freeze_panes = "C1"  # Fija las primeras dos columnas y la fila de encabezado

    # Ajustar ancho de columnas automáticamente
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(1, len(fixed_columns) + len(column_names) + 1):
        dim_holder[get_column_letter(col)] = ws.column_dimensions[get_column_letter(col)]
        dim_holder[get_column_letter(col)].width = 15
    ws.column_dimensions = dim_holder

    # Hoja adicional: Conceptos de Nómina
    ws_conceptos = wb.create_sheet(title="Conceptos de Nómina")

    # Obtener los conceptos desde la base de datos
    usuario = request.session.get('usuario', {})
    idempresa = usuario.get('idempresa')
    conceptos_data = Conceptosdenomina.objects.filter(id_empresa_id=idempresa).values('nombreconcepto', 'codigo').order_by('codigo')

    # Escribir encabezados en la hoja de conceptos
    ws_conceptos.cell(row=1, column=1, value="Nombre del Concepto")
    ws_conceptos.cell(row=1, column=2, value="Código")

    # Aplicar estilo para que los encabezados estén centrados
    ws_conceptos.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws_conceptos.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")

    # Escribir los conceptos en la hoja
    for row_num, concepto in enumerate(conceptos_data, start=2):
        ws_conceptos.cell(row=row_num, column=1, value=concepto['nombreconcepto'])
        # Convertir el código a entero
        ws_conceptos.cell(row=row_num, column=2, value=int(concepto['codigo']))

    # Ajustar ancho de columnas automáticamente en la hoja de conceptos
    dim_holder_conceptos = DimensionHolder(worksheet=ws_conceptos)
    for col in range(1, 3):  # Solo dos columnas (nombreconcepto y codigo)
        dim_holder_conceptos[get_column_letter(col)] = ws_conceptos.column_dimensions[get_column_letter(col)]
        dim_holder_conceptos[get_column_letter(col)].width = 25
    ws_conceptos.column_dimensions = dim_holder_conceptos

    # Guardar el archivo Excel en un HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=payroll_file.xlsx'
    wb.save(response)
    return response



def flat_modal(request):
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']
    
    # Diccionario para almacenar los errores por fila
    errors = [] 
    general_error = []
    
    if request.method == 'POST':
        file = request.FILES.get("file")

        if file:
            file_name = file.name  # Nombre del archivo
            file_size = file.size  # Tamaño del archivo
            
            # Determinar el formato del archivo
            if file_name.endswith('.csv'):
                df = pd.read_csv(file)  # Leer CSV directamente
            elif file_name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file)  # Leer Excel directamente
            else:
                return JsonResponse({"message": "Formato de archivo no soportado"}, status=400)
            
            df = df.fillna(0) # Reemplazar NaN por 0
            # Definir las columnas fijas
            fixed_columns = ['Contrato', 'Nombre']

            # Identificar las columnas dinámicas
            dynamic_columns = [col for col in df.columns if col not in fixed_columns]
    
            # Validar que las columnas fijas existan
            missing_columns = [col for col in fixed_columns if col not in df.columns]
            
            if missing_columns:
                errors.append({'general': f'Faltan las siguientes columnas obligatorias: {", ".join(missing_columns)}.'})
                return JsonResponse({"message": "Ha ocurrido un error ","errors": errors }, status=400)

            conceptos = Conceptosdenomina.objects.filter(id_empresa_id=idempresa).values_list('codigo', flat=True)

            general_error.append(validate_concepts(df,list(conceptos)))
            
            # Validar filas y generar errores
            for index, row in df.iterrows():
                row_errors = [] 

                # Validar las columnas fijas
                idcontrato = row.get('Contrato')
                
                if pd.isnull(idcontrato):
                    row_errors.append('1')
                else : 
                    contrato = Contratos.objects.filter(idcontrato=idcontrato, id_empresa_id=idempresa).first()                
                
                if contrato is None:
                    row_errors.append('3')
                else:
                    if contrato.estadocontrato != 1:
                        row_errors.append('4')
                        
                    if contrato.id_empresa_id != idempresa:
                        row_errors.append('6')
                        
                if row_errors:
                    errors.append({
                        'line': index + 1,  # Fila 1 en Excel corresponde al índice 0 en DataFrame
                        'contract_id': contrato.idcontrato  if contrato is not None else "Contrato no disponible" ,
                        'identification': contrato.idempleado.docidentidad if contrato is not None else "Identificación no disponible",
                        'name':  f"{contrato.idempleado.pnombre} {contrato.idempleado.papellido}" if contrato is not None else "Nombre no disponible" ,  # Suponiendo que no se tiene la columna Nombre
                        'error': "".join([f"<li>{error_messages[err]}</li>" for err in row_errors]) , 
                        'details' : 'detalles' 
                    })
            
            
            if errors:
                return JsonResponse({
                    "message": "¡Se encontraron errores en el archivo!",
                    "errors": errors,
                    "general_error": general_error
                })
            else :
                return JsonResponse({
                    "message": "¡Operación completada con éxito!",

                })

        return JsonResponse({"message": "No se recibió ningún archivo"}, status=400)
    
    return render(request, './payroll/partials/flat_modal.html')



