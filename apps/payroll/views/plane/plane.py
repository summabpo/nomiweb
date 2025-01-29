from django.shortcuts import render
from django.core.files.storage import default_storage
from django.contrib.auth.decorators import login_required
from apps.components.decorators import role_required
import pandas as pd
from openpyxl.utils.exceptions import InvalidFileException
from apps.common.models import Conceptosdenomina, Contratos
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


# Diccionario de mensajes de error con números como claves
error_messages = {
    '1': "Error code 1001: Id contrato inválido o faltante",
    '2': "Error code 1002: Concepto faltante o no es un número entero",
    '3': "Error code 1003: Contrato no encontrado",
    '4': "Error code 1004: El estado del contrato no es válido (debe estar activo)",
    '5': "Error code 1005: Concepto no encontrado",
    '6': "Error code 1006: El contrato no pertenece a la empresa actual",
    '7': "Error code 1007: Fila vacía o incompleta",
    '8': "Error code 1008: El valor debe ser mayor que cero",
    '9': "Error code 1009: Concepto duplicado para este contrato",
    '10': "Error code 1010: Faltan valores críticos (Id contrato, Concepto)",
    '11': "Error code 1012: La cantidad no puede ser negativa",
    '12': "Error code 1013: El contrato está cerrado o inactivo",
    '13': "Error code 1014: El valor no puede ser negativo",
    '15': "Error code 1015: El contrato y concepto ya existen",
    'general': "Error general: Ocurrió un error procesando el archivo",
}

@login_required
@role_required('accountant')
def plane(request, id):
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']
    
    # Diccionario para almacenar los errores por fila
    errors = [] 
    
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        
        if not file.name.endswith('.xlsx'):
            errors.append({'general': 'Solo se aceptan archivos con extensión .xlsx.'})
            return render(request, './payroll/plane.html', {'id': id, 'errors': errors})
        
        try:
            file_name = default_storage.save(file.name, file)
            file_path = default_storage.path(file_name)
            df = pd.read_excel(file_path)
            df.columns = df.columns.str.strip()            
            # Definir las columnas fijas
            fixed_columns = ['Contrato', 'Nombre']

            # Identificar las columnas dinámicas
            dynamic_columns = [col for col in df.columns if pd.notna(col) and col not in fixed_columns]

            print(dynamic_columns)
            # Validar que las columnas fijas existan
            missing_columns = [col for col in fixed_columns if col not in df.columns]
            
            if missing_columns:
                errors.append({'general': f'Faltan las siguientes columnas obligatorias: {", ".join(missing_columns)}.'})
                return render(request, './payroll/plane.html', {'id': id, 'errors': errors})

            # Validar filas y generar errores
            for index, row in df.iterrows():
                row_errors = [] 

                # Validar las columnas fijas
                idcontrato = row.get('Contrato')
                
                if pd.isnull(idcontrato):
                    row_errors.append('1')
                else : 
                    contrato = Contratos.objects.filter(idcontrato=idcontrato, id_empresa_id=idempresa).first()                
                
                if contrato is None:
                    row_errors.append('3')
                else:
                    if contrato.estadocontrato != '1':
                        row_errors.append('4')
                        
                    if contrato.id_empresa_id != idempresa:
                        row_errors.append('6')

                for col in dynamic_columns:
                    
                    concepto = Conceptosdenomina.objects.filter(codigo=col, id_empresa_id=idempresa).first()
                    if concepto is None:
                        row_errors.append('5')
                
                # Agregar errores por fila si existen
                if row_errors:
                    errors.append({
                        'line': index + 1,  # Fila 1 en Excel corresponde al índice 0 en DataFrame
                        'contract_id': contrato.idcontrato  if contrato is not None else "Contrato no disponible" ,
                        'identification': contrato.idempleado.docidentidad if contrato is not None else "Identificación no disponible",
                        'name':  f"{contrato.idempleado.pnombre} {contrato.idempleado.papellido}" if contrato is not None else "Nombre no disponible" ,  # Suponiendo que no se tiene la columna Nombre
                        'error': "".join([f"<li>{error_messages[err]}</li>" for err in row_errors]) , 
                        'details' : 'detalles' 
                    })
                    
                    
            # Borrar el archivo una vez procesado
            default_storage.delete(file_name)

        except Exception as e:
            errors.append({'general': f'Ocurrió un error procesando el archivo: {str(e)}'})
    print(errors)
    return render(request, './payroll/plane.html', {
        'id': id,
        'errors': errors,
    })





def document(request):
    # Crear el libro de Excel
    wb = Workbook()
    
    # Hoja principal: Datos
    ws = wb.active
    ws.title = "Datos"

    # Definir las columnas fijas
    fixed_columns = ["Contrato", "Nombre"]
    # Definir las columnas dinámicas
    column_names = ["15", "10", "11"]

    # Escribir las columnas fijas
    for col_num, col_name in enumerate(fixed_columns, 1):
        ws.cell(row=1, column=col_num, value=col_name)

    # Escribir las columnas dinámicas
    for col_num, col_name in enumerate(column_names, len(fixed_columns) + 1):
        ws.cell(row=1, column=col_num, value=col_name)

    # Aplicar estilo para que los encabezados estén centrados
    for col in range(1, len(fixed_columns) + len(column_names) + 1):
        ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

    # Fijar las primeras dos columnas
    ws.freeze_panes = "C1"  # Fija las primeras dos columnas y la fila de encabezado

    # Ajustar ancho de columnas automáticamente
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(1, len(fixed_columns) + len(column_names) + 1):
        dim_holder[get_column_letter(col)] = ws.column_dimensions[get_column_letter(col)]
        dim_holder[get_column_letter(col)].width = 15
    ws.column_dimensions = dim_holder

    # Hoja adicional: Conceptos de Nómina
    ws_conceptos = wb.create_sheet(title="Conceptos de Nómina")

    # Obtener los conceptos desde la base de datos
    usuario = request.session.get('usuario', {})
    idempresa = usuario.get('idempresa')
    conceptos_data = Conceptosdenomina.objects.filter(id_empresa_id=idempresa).values('nombreconcepto', 'codigo').order_by('codigo')

    # Escribir encabezados en la hoja de conceptos
    ws_conceptos.cell(row=1, column=1, value="Nombre del Concepto")
    ws_conceptos.cell(row=1, column=2, value="Código")

    # Aplicar estilo para que los encabezados estén centrados
    ws_conceptos.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws_conceptos.cell(row=1, column=2).alignment = Alignment(horizontal="center", vertical="center")

    # Escribir los conceptos en la hoja
    for row_num, concepto in enumerate(conceptos_data, start=2):
        ws_conceptos.cell(row=row_num, column=1, value=concepto['nombreconcepto'])
        ws_conceptos.cell(row=row_num, column=2, value=int(concepto['codigo']))

    # Ajustar ancho de columnas automáticamente en la hoja de conceptos
    dim_holder_conceptos = DimensionHolder(worksheet=ws_conceptos)
    for col in range(1, 3):  # Solo dos columnas (nombreconcepto y codigo)
        dim_holder_conceptos[get_column_letter(col)] = ws_conceptos.column_dimensions[get_column_letter(col)]
        dim_holder_conceptos[get_column_letter(col)].width = 25
    ws_conceptos.column_dimensions = dim_holder_conceptos

    # Guardar el archivo Excel en un HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=plane_nomina.xlsx'
    wb.save(response)
    return response
