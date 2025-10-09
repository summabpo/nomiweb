import openpyxl
from openpyxl.styles import NamedStyle
from io import BytesIO
from django.http import HttpResponse
from apps.common.models import Contratos
from django.db.models import Sum


FormaPago = (
    ('', '----------'),
    ('1', 'Abono a cuenta'),
    ('2', 'Cheque'),
    ('3', 'Efectivo'),
    ('4', 'Transferencia electrónica'),
)



""" 
original :
'Documento', 'Nombre', 'Fecha Inicio Contrato', 'Cargo', 'Salario',
        'Centro de Costos', 'Tipo de Contrato', 'Tarifa ARL', 'Fecha Fin Contrato',
        'Tipo Nomina', 'Banco Cuenta', 'Cuenta Nomina', 'Tipo Cuenta Nomina', 'EPS', 'Pension',
        'Caja Compensacion', 'Ciudad Contratacion ', 'Fondo Cesantias',
        'Forma Pago', 'Tipo Salario', 'Modelo', 'Departamento',
        'Ciudad'
quitados 
'Fondo Cesantias',

"""

def generate_contract_excel(idempresa):
    """
    Genera un archivo Excel con los detalles de los contratos activos de los empleados de una empresa.
    Limpia los textos 'no data' antes de exportar.
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import NamedStyle

    def clean(value):
        """Convierte 'no data' o valores None en vacío."""
        if isinstance(value, str) and value.strip().lower() == "no data":
            return ""
        return value or ""

    # Crear archivo en memoria
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract Report"

    # Encabezados
    headers = [
        'Documento', 'Nombre', 'Fecha Inicio Contrato', 'Cargo', 'Salario',
        'Centro de Costos', 'Tipo de Contrato', 'Tarifa ARL', 'Fecha Fin Contrato',
        'Tipo Nomina', 'Banco Cuenta', 'Cuenta Nomina', 'Tipo Cuenta Nomina', 'EPS', 'Pension',
        'Caja Compensacion', 'Ciudad Contratacion ', 
        'Forma Pago', 'Tipo Salario', 'Modelo', 'Departamento', 'ID Contrato'
    ]
    ws.append(headers)

    decimal_style = NamedStyle(name='decimal_style', number_format='0.00')

    # Consulta optimizada
    contratos_empleados = Contratos.objects.filter(
        estadocontrato=1, id_empresa=idempresa
    ).only(
        'idempleado__docidentidad', 'idempleado__papellido', 'idempleado__pnombre',
        'idempleado__snombre', 'fechainiciocontrato', 'cargo', 'salario', 'idcosto__nomcosto',
        'tipocontrato__tipocontrato', 'centrotrabajo__tarifaarl', 'fechafincontrato',
        'tiponomina', 'bancocuenta', 'cuentanomina', 'tipocuentanomina', 'codeps',
        'codafp', 'codccf', 'ciudadcontratacion__ciudad', 'formapago',
        'tiposalario__tiposalario', 'jornada', 'idmodelo__nombremodelo', 'idcontrato'
    )

    forma_pago_dict = dict(FormaPago)

    for data in contratos_empleados:
        nombre = " ".join(filter(None, [
            clean(data.idempleado.papellido),
            clean(data.idempleado.pnombre),
            clean(data.idempleado.snombre),
        ])).strip()

        row = [
            clean(data.idempleado.docidentidad),
            nombre,
            clean(data.fechainiciocontrato),
            clean(data.cargo.nombrecargo),
            clean(data.salario),
            clean(data.idcosto.nomcosto),
            clean(data.tipocontrato.tipocontrato),
            clean(data.centrotrabajo.tarifaarl),
            clean(data.fechafincontrato),
            clean(data.tiponomina.tipodenomina),
            clean(data.bancocuenta.nombanco if data.bancocuenta else ''),
            clean(data.cuentanomina),
            clean(data.tipocuentanomina),
            clean(data.codeps.entidad),
            clean(data.codafp.entidad),
            clean(data.codccf.entidad),
            clean(data.ciudadcontratacion.ciudad),
            clean(forma_pago_dict.get(data.formapago, '')),
            clean(data.tiposalario.tiposalario),
            clean(data.jornada),
            clean(data.idmodelo.nombremodelo),
            clean(data.idcontrato),
        ]
        ws.append(row)

    # Formato decimal
    for cell in ws['E']:
        cell.style = decimal_style

    # Ajustar ancho columnas
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    wb.save(output)
    output.seek(0)
    return output.getvalue()



def generate_contract_start_excel(idempresa: int) -> bytes:
    """
    Genera un archivo Excel (.xlsx) con un resumen de los contratos activos 
    de los empleados pertenecientes a una empresa específica.
    Limpia cualquier texto 'no data' en los campos exportados.
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import NamedStyle

    def clean(value):
        """Convierte 'no data' o valores None en vacío."""
        if isinstance(value, str) and value.strip().lower() == "no data":
            return ""
        return value or ""

    # Crear archivo Excel en memoria
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contract Report"

    # Definir encabezados
    headers = [
        'Documento', 'Nombre', 'Fecha Inicio Contrato', 'Cargo', 'Salario',
        'Centro de Costos', 'Tipo de Contrato', 'Tarifa ARL', 'ID Contrato'
    ]
    ws.append(headers)

    # Estilo para formato decimal (salario)
    decimal_style = NamedStyle(name='decimal_style', number_format='#,##0.00')

    # Consulta optimizada
    contratos_empleados = (
        Contratos.objects.filter(estadocontrato=1, id_empresa=idempresa)
        .select_related('idempleado', 'idcosto', 'tipocontrato', 'centrotrabajo', 'cargo')
        .only(
            'idempleado__docidentidad',
            'idempleado__papellido',
            'idempleado__pnombre',
            'idempleado__snombre',
            'fechainiciocontrato',
            'cargo__nombrecargo',
            'salario',
            'idcosto__nomcosto',
            'tipocontrato__tipocontrato',
            'centrotrabajo__tarifaarl',
            'idcontrato',
        )
    )

    # Insertar datos
    for contrato in contratos_empleados:
        nombre = " ".join(filter(None, [
            clean(contrato.idempleado.papellido),
            clean(contrato.idempleado.pnombre),
            clean(contrato.idempleado.snombre)
        ])).strip()

        row = [
            clean(contrato.idempleado.docidentidad),
            nombre,
            clean(contrato.fechainiciocontrato.strftime("%Y-%m-%d") if contrato.fechainiciocontrato else ""),
            clean(contrato.cargo.nombrecargo if contrato.cargo else ""),
            clean(contrato.salario),
            clean(contrato.idcosto.nomcosto if contrato.idcosto else ""),
            clean(contrato.tipocontrato.tipocontrato if contrato.tipocontrato else ""),
            clean(contrato.centrotrabajo.tarifaarl if contrato.centrotrabajo else ""),
            clean(contrato.idcontrato),
        ]
        ws.append(row)

    # Aplicar formato decimal a columna E (Salario)
    for cell in ws['E'][1:]:  # omitir encabezado
        cell.style = decimal_style

    # Ajustar ancho de columnas automáticamente
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    # Guardar en memoria
    wb.save(output)
    output.seek(0)

    return output.getvalue()


