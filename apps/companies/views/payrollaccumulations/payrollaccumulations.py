from django.shortcuts import render, redirect
from apps.common.models import Nomina , Conceptosdenomina
from apps.companies.forms.ReportFilterForm import ReportFilterForm , ReportFilter2Form
from django.contrib import messages
from django.http import HttpResponse
from apps.components.generate_employee_excel import generate_employee_excel
from apps.components.humani import format_value
from django.http import JsonResponse
from .parse_dates import parse_dates
from django.db.models import Sum, Q
from apps.components.decorators import  role_required
from django.contrib.auth.decorators import login_required
from openpyxl.styles import Font
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment


@login_required
@role_required('company', 'accountant')
def payrollaccumulations(request):

    acumulados = {}
    compects = []
    usuario = request.session.get('usuario', {})
    idempresa = usuario.get('idempresa')

    form = ReportFilterForm(idempresa=idempresa)

    # 🔹 Limpieza de valores
    def clean_value(value):
        if not value:
            return ''
        return str(value).replace('no data', '').strip()

    MES_ORDER = {
        'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
        'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
        'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
    }

    if request.method == 'POST':
        form = ReportFilterForm(request.POST, idempresa=idempresa)

        if form.is_valid():
            employee = form.cleaned_data['employee']
            cost_center = form.cleaned_data['cost_center']
            city = form.cleaned_data['city']
            year_init = form.cleaned_data.get('year_init')
            mst_init = form.cleaned_data.get('mst_init')
            year_end = form.cleaned_data.get('year_end')
            mst_end = form.cleaned_data.get('mst_end')

            # 🔥 Construcción de filtros eficientes
            filters = Q(idnomina__id_empresa_id=idempresa)

            if employee:
                filters &= Q(idcontrato__idempleado__idempleado=employee)

            if cost_center:
                filters &= Q(idcontrato__idcosto=cost_center)

            if city:
                filters &= Q(idcontrato__idsede=city)

            # 🔹 Filtro por rango de fechas optimizado
            if year_init and mst_init and year_end and mst_end:
                try:
                    year_init = int(year_init)
                    year_end = int(year_end)
                except ValueError:
                    messages.error(request, "Los años deben ser números válidos.")
                    year_init = year_end = None

                inicio_num = MES_ORDER.get(mst_init.upper())
                fin_num = MES_ORDER.get(mst_end.upper())

                if inicio_num and fin_num and year_init and year_end:

                    filters &= (
                        Q(
                            idnomina__anoacumular__ano__gt=year_init,
                            idnomina__anoacumular__ano__lt=year_end
                        )
                        |
                        Q(
                            idnomina__anoacumular__ano=year_init,
                            idnomina__mesacumular__in=[
                                mes for mes, num in MES_ORDER.items() if num >= inicio_num
                            ]
                        )
                        |
                        Q(
                            idnomina__anoacumular__ano=year_end,
                            idnomina__mesacumular__in=[
                                mes for mes, num in MES_ORDER.items() if num <= fin_num
                            ]
                        )
                    )

            # 🔥 QUERY OPTIMIZADA (GROUP BY EN DB)
            nominas = (
                Nomina.objects
                .filter(filters)
                .values(
                    'idcontrato__idempleado__docidentidad',
                    'idcontrato__idempleado__papellido',
                    'idcontrato__idempleado__sapellido',
                    'idcontrato__idempleado__pnombre',
                    'idcontrato__idempleado__snombre',
                    'idcontrato__idcontrato',
                    'idconcepto__codigo',
                    'idconcepto__nombreconcepto',
                )
                .annotate(
                    total_cantidad=Sum('cantidad'),
                    total_valor=Sum('valor')
                )
                .order_by(
                    'idcontrato__idempleado__docidentidad',
                    'idconcepto__codigo'
                )
            )

            # 🔹 Construcción del resultado (ligero)
            for row in nominas:
                doc = row['idcontrato__idempleado__docidentidad']

                nombre = " ".join(filter(None, [
                    clean_value(row['idcontrato__idempleado__papellido']),
                    clean_value(row['idcontrato__idempleado__sapellido']),
                    clean_value(row['idcontrato__idempleado__pnombre']),
                    clean_value(row['idcontrato__idempleado__snombre']),
                ]))

                if doc not in acumulados:
                    acumulados[doc] = {
                        'documento': doc,
                        'empleado': nombre,
                        'contrato': row['idcontrato__idcontrato'],
                        'data': []
                    }

                acumulados[doc]['data'].append({
                    "idconcepto": row['idconcepto__codigo'],
                    "nombreconcepto": clean_value(row['idconcepto__nombreconcepto']),
                    "cantidad": row['total_cantidad'] or 0,
                    "valor": format_value(row['total_valor'] or 0),
                })

            compects = list(acumulados.values())

        else:
            for error in form.errors.values():
                for e in error:
                    messages.error(request, e)

    return render(request, 'companies/payrollaccumulations.html', {
        'compects': compects,
        'form': form,
    })
    
    
    
@login_required
@role_required('company','accountant')
def payrollaccumulations2(request):
    """
    Vista para consultar los acumulados de nómina de los empleados de una empresa.

    Esta vista permite a usuarios con el rol 'company' o 'accountant' aplicar filtros como empleado, centro de costos, ciudad,
    año y mes para generar una lista de conceptos acumulados por empleado en el periodo seleccionado. La información es agrupada
    por empleado y concepto, sumando cantidad y valor. También se formatea el valor antes de enviarlo al template.

    Parameters
    ----------
    request : HttpRequest
        Objeto de solicitud HTTP que puede contener parámetros POST con los filtros seleccionados por el usuario.

    Returns
    -------
    HttpResponse
        Devuelve una respuesta renderizada con el template 'companies/payrollaccumulations.html', que incluye el formulario de filtros
        y los datos acumulados por empleado y concepto.

    Notes
    -----
    El usuario debe estar autenticado y tener el rol 'company' o 'accountant' para acceder a esta vista.
    En caso de errores de validación en el formulario, se muestran mensajes de error al usuario.
    """

    acumulados = {}
    compects = []

    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']

    form = ReportFilter2Form(idempresa=idempresa)

    # --- Función auxiliar ---
    def clean_value(value):
        if not value:
            return ''
        return str(value).replace('no data', '').strip()

    # Orden de meses
    MES_ORDER = {
        'ENERO': 1,
        'FEBRERO': 2,
        'MARZO': 3,
        'ABRIL': 4,
        'MAYO': 5,
        'JUNIO': 6,
        'JULIO': 7,
        'AGOSTO': 8,
        'SEPTIEMBRE': 9,
        'OCTUBRE': 10,
        'NOVIEMBRE': 11,
        'DICIEMBRE': 12
    }

    if request.method == 'POST':
        form = ReportFilter2Form(request.POST, idempresa=idempresa)

        if form.is_valid():
            year_init = form.cleaned_data.get('year_init')
            mst_init = form.cleaned_data.get('mst_init')
            inicio_num = MES_ORDER.get(mst_init.upper())

            # -------------------------
            # FILTRO POR AÑO Y MES
            # -------------------------
            if year_init and mst_init:
                try:
                    year_init = int(year_init)
                except ValueError:
                    messages.error(request, "El año debe ser un número válido.")
                    year_init = None

                inicio_num = MES_ORDER.get(mst_init.upper())

                if inicio_num and year_init:
                    nominas = Nomina.objects.filter(
                        idnomina__id_empresa_id=idempresa ,
                        idnomina__anoacumular__ano=year_init,
                        idnomina__mesacumular=mst_init.upper()
                    ).select_related(
                        'idcontrato',
                        'idcontrato__idempleado',
                        'idconcepto'
                    ).order_by('-idregistronom')

            # -------------------------
            # ACUMULADOS
            # -------------------------
            for data in nominas:
                contrato = data.idcontrato
                empleado = contrato.idempleado
                doc = empleado.docidentidad

                if doc not in acumulados:
                    nombre_completo = " ".join(filter(None, [
                        clean_value(empleado.papellido),
                        clean_value(empleado.sapellido),
                        clean_value(empleado.pnombre),
                        clean_value(empleado.snombre),
                    ]))

                    acumulados[doc] = {
                        'documento': doc,
                        'empleado': nombre_completo,
                        'contrato': contrato.idcontrato,
                        'data': []
                    }

                conceptos = acumulados[doc]['data']

                concepto_existente = next(
                    (c for c in conceptos if c["idconcepto"] == data.idconcepto.codigo),
                    None
                )

                if concepto_existente:
                    concepto_existente["cantidad"] += data.cantidad or 0
                    concepto_existente["valor"] += data.valor or 0
                else:
                    conceptos.append({
                        "idconcepto": data.idconcepto.codigo,
                        "nombreconcepto": clean_value(data.idconcepto.nombreconcepto),
                        "cantidad": data.cantidad or 0,
                        "valor": data.valor or 0,
                    })

            # -------------------------
            # FORMATEAR VALORES
            # -------------------------
            for datos in acumulados.values():
                for concepto in datos['data']:
                    concepto['valor'] = format_value(concepto['valor'])

            compects = list(acumulados.values())

        else:
            for error in form.errors.values():
                for e in error:
                    messages.error(request, e)

    return render(request, 'companies/payrollaccumulations2.html', {
        'compects': compects,
        'form': form,
    })


    
@login_required
@role_required('company','accountant')  
def descargar_excel_empleados(request):
    """
    Vista para descargar en Excel los acumulados de nómina por empleado.

    Esta vista recibe una solicitud POST con parámetros de filtrado (fechas, empleado, centro de costos, ciudad)
    y genera un archivo Excel con los conceptos de nómina acumulados por empleado dentro del rango dado.

    Parameters
    ----------
    request : HttpRequest
        Objeto de solicitud HTTP que debe contener datos POST con los filtros de búsqueda.

    Returns
    -------
    HttpResponse
        Devuelve un archivo Excel generado con los datos acumulados si la solicitud es POST y válida.

    JsonResponse
        Si la solicitud no es POST, devuelve una respuesta de error con estado HTTP 405.

    Notes
    -----
    La vista requiere autenticación y el rol adecuado. Utiliza la función `parse_dates` para descomponer las fechas 
    y `generate_employee_excel` para generar el archivo de Excel.
    """
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']

    def clean_value(value):
        return str(value).replace('no data', '').strip() if value else ''

    if request.method != 'POST':
        return JsonResponse({'error': 'Falla en la creacion de Documento'}, status=405)

    # -------------------------
    # Filtros
    # -------------------------
    filtros = {}
    post = request.POST

    if post.get('employee'):
        filtros['idempleado__idempleado'] = post['employee']
    if post.get('cost_center'):
        filtros['idcontrato__idcosto'] = post['cost_center']
    if post.get('city'):
        filtros['idcontrato__idsede'] = post['city']
    if post.get('start_date'):
        filtros['idnomina__fechainicial__gte'] = post['start_date']
    if post.get('end_date'):
        filtros['idnomina__fechafinal__lte'] = post['end_date']
    if idempresa:
        filtros['idcontrato__id_empresa'] = idempresa
    # -------------------------
    # Query optimizada
    # -------------------------
    nominas = (
        Nomina.objects
        .filter(**filtros)
        .select_related(
            'idcontrato',
            'idcontrato__idempleado',
            'idconcepto'
        )
        .order_by('idconcepto__idconcepto')
    )

    acumulados = {}

    # -------------------------
    # Procesamiento
    # -------------------------
    for data in nominas:
        contrato = data.idcontrato
        empleado = contrato.idempleado
        doc = empleado.docidentidad

        if doc not in acumulados:
            nombre_completo = " ".join(filter(None, [
                clean_value(empleado.papellido),
                clean_value(empleado.sapellido),
                clean_value(empleado.pnombre),
                clean_value(empleado.snombre),
            ]))

            acumulados[doc] = {
                'documento': doc,
                'Empleado': f"{nombre_completo} - {doc} - {contrato.idcontrato}",
                'data': {},
                'total': 0,
                'id': contrato.idcontrato,
            }

        concepto_id = data.idconcepto.idconcepto

        # 🔥 cambio clave: dict en vez de lista (O(1))
        conceptos = acumulados[doc]['data']

        if concepto_id not in conceptos:
            conceptos[concepto_id] = {
                "idconcepto": concepto_id,
                "nombreconcepto": data.idconcepto.nombreconcepto,
                "cantidad": 0,
                "valor": 0,
            }

        conceptos[concepto_id]["cantidad"] += data.cantidad
        conceptos[concepto_id]["valor"] += data.valor

        acumulados[doc]['total'] += data.valor

    # Convertir dict de conceptos a lista (para tu Excel)
    for emp in acumulados.values():
        emp['data'] = list(emp['data'].values())

    # -------------------------
    # Fechas y Excel
    # -------------------------
    start_year, start_month, end_year, end_month = parse_dates(
        post.get('start_date'),
        post.get('end_date')
    )

    output = generate_employee_excel(
        acumulados,
        start_month,
        start_year,
        end_month,
        end_year
    )

    return HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="empleado_info.xlsx"'}
    )




@login_required
@role_required('company', 'accountant')  
def descargar_excel_empleados_2(request):
    
    if request.method != 'POST':
        return JsonResponse({'error': 'Método no permitido'}, status=405)

    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']

    def clean_value(value):
        return str(value).replace('no data', '').strip() if value else ''

    post = request.POST

    year_init = post.get('year_init')
    mst_init = post.get('mst_init')

    # -------------------------
    # Query base
    # -------------------------
    nominas = Nomina.objects.filter(
        idcontrato__id_empresa=idempresa,
    ).select_related(
        'idcontrato',
        'idcontrato__idempleado',
        'idconcepto',
        'idnomina__anoacumular'
    ).order_by('idconcepto__codigo')

    # -------------------------
    # Filtro por año/mes (ACUMULADO)
    # -------------------------
    MES_ORDER = {
        'ENERO': 1,'FEBRERO': 2,'MARZO': 3,'ABRIL': 4,'MAYO': 5,'JUNIO': 6,
        'JULIO': 7,'AGOSTO': 8,'SEPTIEMBRE': 9,'OCTUBRE': 10,'NOVIEMBRE': 11,'DICIEMBRE': 12
    }

    if year_init and mst_init:
        try:
            year_init = int(year_init)
        except ValueError:
            year_init = None

        inicio_num = MES_ORDER.get(mst_init.upper())

        if inicio_num and year_init:
            nominas = nominas.filter(
                idnomina__anoacumular__ano=year_init,
                idnomina__mesacumular=mst_init.upper()
            )



    # for n in nominas:
    #     #if n.idconcepto.codigo == 31 or n.idconcepto.codigo == 83 :

    #     print(
    #         f"Empleado: {n.idcontrato} | "
    #         f"Contrato: {n.idcontrato.idcontrato} | "
    #         f"Concepto: {n.idconcepto.nombreconcepto} | "
    #         f"Código: {n.idconcepto.codigo} | "
    #         f"mes: {n.idnomina.mesacumular} | "
    #         f"año: {n.idnomina.anoacumular.ano} | "
    #         f"valor: {n.valor} | "
    #         f"Cantidad: {n.cantidad}"
    #     )


    # -------------------------
    # Conceptos Ley 1393
    # -------------------------
    conceptos_1393_ids = set(
        Conceptosdenomina.objects.filter(
            indicador__nombre="base1393",
            id_empresa = idempresa,
        ).values_list("codigo", flat=True)
    )

    # -------------------------
    # Acumulados
    # -------------------------
    acumulados = {}
    especiales = {1, 2, 4, 34, 25, 26, 27, 28, 31, 83, 82}

    for data in nominas:
        contrato = data.idcontrato
        empleado = contrato.idempleado
        doc = empleado.docidentidad

        if doc not in acumulados:
            nombre_completo = " ".join(filter(None, [
                clean_value(empleado.papellido),
                clean_value(empleado.sapellido),
                clean_value(empleado.pnombre),
                clean_value(empleado.snombre),
            ]))

            acumulados[doc] = {
                "Documento": doc, 
                "Empleado": nombre_completo,
                "Cargo": contrato.cargo.nombrecargo if contrato.cargo else "",
                "Costo": contrato.idcosto.nomcosto if contrato.idcosto else "",
                "Fecha_inicio": contrato.fechainiciocontrato,
                "Fecha_retiro": contrato.fechafincontrato,
                "Salario": contrato.salario,
                "conceptos": {},  
            }

        concepto_nombre = f"{data.idconcepto.nombreconcepto} ({data.idconcepto.codigo})"
        codigo = data.idconcepto.codigo
        conceptos = acumulados[doc]['conceptos']

        if concepto_nombre not in conceptos:
            conceptos[concepto_nombre] = {"valor": 0, "cantidad": 0}

        # ✅ Acumular valor SIEMPRE
        conceptos[concepto_nombre]["valor"] += data.valor

        # ✅ Acumular cantidad SOLO si es especial
        if codigo in especiales:
            conceptos[concepto_nombre]["cantidad"] += data.cantidad

    # -------------------------
    # Columnas dinámicas
    # -------------------------
    conceptos_ordenados = []
    otros = []

    for data in nominas:
        nombre = f"{data.idconcepto.nombreconcepto} ({data.idconcepto.codigo})"
        codigo = data.idconcepto.codigo

        if codigo in especiales and nombre not in conceptos_ordenados:
            conceptos_ordenados.append(nombre)
        elif codigo not in especiales and nombre not in otros:
            otros.append(nombre)

    conceptos_ordenados.extend(otros)

    # -------------------------
    # Mapa concepto -> código
    # -------------------------
    concepto_codigo_map = {}

    for data in nominas:
        nombre = f"{data.idconcepto.nombreconcepto} ({data.idconcepto.codigo})"
        concepto_codigo_map[nombre] = data.idconcepto.codigo

    # -------------------------
    # Separar conceptos positivos y negativos
    # -------------------------
    conceptos_pos = []
    conceptos_neg = []

    for concepto in conceptos_ordenados:
        total = 0
        for emp in acumulados.values():
            total += emp['conceptos'].get(concepto, {}).get("valor", 0)

        if total >= 0:
            conceptos_pos.append(concepto)
        else:
            conceptos_neg.append(concepto)

    # -------------------------
    # Crear Excel
    # -------------------------

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # azul oscuro
    header_font = Font(bold=True, color="FFFFFF")  # blanco
    header_alignment = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina"

    headers = ["Documento", "Empleado","Cargo","Centro de Costo","Fecha Ingreso","Fecha Retiro","Sueldo"]

    # -------- POSITIVOS --------
    for concepto in conceptos_pos:
        headers.append(f"{concepto} - Valor")
        if concepto_codigo_map.get(concepto) in especiales:
            headers.append(f"{concepto} - Cantidad")

    headers.append("TOTAL DEVENGADOS")

    # -------- NEGATIVOS --------
    for concepto in conceptos_neg:
        headers.append(f"{concepto} - Valor")
        if concepto_codigo_map.get(concepto) in especiales:
            headers.append(f"{concepto} - Cantidad")

    headers.append("TOTAL DEDUCCIONES")

    # -------- TOTAL FINAL --------
    headers.append("TOTAL NETO")

    headers.append("")  # columna vacía (espacio)
    headers.append("LEY 1393")

    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # -------------------------
    # Inicializar totales por columna
    # -------------------------
    totales_columnas = [0] * len(headers)

    # -------------------------
    # Data
    # -------------------------
    for emp in acumulados.values():
        fila = [
            emp["Documento"],
            emp["Empleado"],
            emp["Cargo"],
            emp["Costo"],
            emp["Fecha_inicio"],
            emp["Fecha_retiro"],
            emp["Salario"]
        ]

        total_devengados = 0
        total_deducciones = 0
        valor_1393 = 0
        aux_1393 = 0
        list_aux = {1,4}

        # -------- POSITIVOS --------
        for concepto in conceptos_pos:
            valor = emp['conceptos'].get(concepto, {}).get("valor", 0)
            fila.append(valor)
            total_devengados += valor

            codigo = concepto_codigo_map.get(concepto)

            if codigo in list_aux:
                aux_1393 += valor

            if codigo in conceptos_1393_ids:
                valor_1393 += valor

            if concepto_codigo_map.get(concepto) in especiales:
                cantidad = emp['conceptos'].get(concepto, {}).get("cantidad", 0)
                fila.append(cantidad)

        fila.append(total_devengados)

        # -------- NEGATIVOS --------
        for concepto in conceptos_neg:
            valor = emp['conceptos'].get(concepto, {}).get("valor", 0)
            fila.append(valor)
            total_deducciones += valor

            codigo = concepto_codigo_map.get(concepto)

            if codigo in conceptos_1393_ids:
                valor_1393 += valor

            if concepto_codigo_map.get(concepto) in especiales:
                cantidad = emp['conceptos'].get(concepto, {}).get("cantidad", 0)
                fila.append(cantidad)

        fila.append(total_deducciones)

        # -------- TOTAL FINAL --------
        fila.append(total_devengados + total_deducciones)

        fila.append("")  # columna vacía

        def calcular_1393(valor_1393, aux_1393):
            base = valor_1393 + aux_1393
            return max(0, valor_1393 - (base * 0.4))

        fila.append(calcular_1393(valor_1393, aux_1393))

        # -------------------------
        # Acumular totales por columna
        # -------------------------
        for i, val in enumerate(fila):
            try:
                totales_columnas[i] += float(val)
            except (TypeError, ValueError):
                pass

        ws.append(fila)

    # -------------------------
    # Fila de totales (negrita)
    # -------------------------
    fila_totales = []

    for i, val in enumerate(totales_columnas):
        if i == 1:
            fila_totales.append("TOTALES")
        elif i < 7:
            fila_totales.append("")
        else:
            fila_totales.append(val)

    ws.append(fila_totales)

    # Aplicar negrita
    bold_font = Font(bold=True)
    for cell in ws[ws.max_row]:
        cell.font = bold_font

    # -------------------------
    # Auto width
    # -------------------------
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # -------------------------
    # Output
    # -------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # -------------------------
    # Nombre dinámico del archivo
    # -------------------------
    mes = mst_init.lower() if mst_init else "todos"
    anio = year_init if year_init else "todos"

    nombre_archivo = f"acumulados_{mes}_{anio}.xlsx"

    return HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre_archivo}"'}
    )




