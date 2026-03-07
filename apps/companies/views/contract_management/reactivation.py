from django.shortcuts import render, redirect, get_object_or_404
from apps.common.models  import Contratos , Contratosemp , Ciudades , Cargos , Entidadessegsocial ,Costos , Bancos
from apps.components.decorators import custom_login_required ,custom_permission
from django.db.models import OuterRef, Exists, Subquery
from apps.components.decorators import  role_required
from django.contrib.auth.decorators import login_required
from apps.companies.forms.ReactivationForm import ReactivationForm
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


ModalidadSalario = (
    ('1', 'Variable'),
    ('2', 'Fijo'),
    ('3', 'Mixto'),
)

FormaPago = (
    ('1', 'Abono a cuenta'),
    ('2', 'Cheque'),
    ('3', 'Efectivo'),
    ('4', 'Transferencia electrónica'),
)


Cercania = (
    (True, 'Si'),
    (False, 'No'),
)

TipoCcuenta = [
    ('ahorros', 'Ahorros'),
    ('corriente', 'Corriente'),
]


@login_required
@role_required('company', 'accountant')
def reactivation(request):  
    
    # --- Función auxiliar de limpieza ---
    def clean_value(value):
        """
        Limpia valores tipo texto que indiquen falta de datos.
        Ejemplo: 'no data', 'sin dato', 'n/a', 'none', 'ninguno'
        """
        if isinstance(value, str):
            if value.strip().lower() in ["no data", "sin dato", "n/a", "none", "ninguno"]:
                return ""
        return value

    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']

    # Contratos activos
    contratos_activos = Contratos.objects.filter(
        idempleado=OuterRef('pk'),
        estadocontrato=1,
        id_empresa_id=idempresa
    )

    # Cualquier contrato (histórico)
    contratos_historicos = Contratos.objects.filter(
        idempleado=OuterRef('pk'),
        id_empresa_id=idempresa
    )

    # Último contrato
    ultimo_contrato = contratos_historicos.order_by('-idcontrato')

    empleados = Contratosemp.objects.filter(
        id_empresa_id=idempresa
    ).annotate(
        tiene_contrato_activo=Exists(contratos_activos),
        tiene_contrato=Exists(contratos_historicos),  # 👈 NUEVO
        fechainiciocontrato=Subquery(ultimo_contrato.values('fechainiciocontrato')[:1]),
        fechafincontrato=Subquery(ultimo_contrato.values('fechafincontrato')[:1]),
        salario=Subquery(ultimo_contrato.values('salario')[:1]),
        cargo=Subquery(ultimo_contrato.values('cargo__nombrecargo')[:1]),
        tipocontrato=Subquery(ultimo_contrato.values('tipocontrato__tipocontrato')[:1]),
        centrocostos=Subquery(ultimo_contrato.values('idcosto__nomcosto')[:1]),
        idcontrato=Subquery(ultimo_contrato.values('idcontrato')[:1]),
    ).filter(
        tiene_contrato_activo=False,   # No activo
        tiene_contrato=True            # Pero sí tuvo contrato
    )

    # Nombre completo
    for e in empleados:
        e.nombre = clean_value(
            f"{e.pnombre or ''} {e.snombre or ''} {e.papellido or ''} {e.sapellido or ''}".strip()
        )
        e.cargo = clean_value(e.cargo)
        e.tipocontrato = clean_value(e.tipocontrato)
        e.centrocostos = clean_value(e.centrocostos)


    return render(request, './companies/reactivation.html', { 'empleados': empleados, 'user': request.user })



@login_required
@role_required('company', 'accountant')
def reactivation_modal(request,id):
    contrato = Contratos.objects.get(idcontrato = id)
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']
    idc = contrato.idcontrato
    data = {
        'name': contrato.idempleado.idempleado,
        'payrollType': contrato.tiponomina.idtiponomina if contrato.tiponomina else '',
        'position': contrato.cargo.idcargo if contrato.cargo else '',
        'workLocation': contrato.ciudadcontratacion.idciudad if contrato.ciudadcontratacion else '',
        'contractType': contrato.tipocontrato.idtipocontrato if contrato.tipocontrato else '',
        'contractModel': contrato.idmodelo.idmodelo if contrato.idmodelo else '',
        'salary': "{:,.0f}".format(contrato.salario).replace(',', '.') if contrato.salario else '',
        'salaryType': contrato.tiposalario.idtiposalario if contrato.tiposalario else '',
        'paymentMethod': contrato.formapago or '',
        'salaryMode': contrato.salariovariable or '',
        'bankAccount': contrato.bancocuenta.idbanco if contrato.bancocuenta else '',
        'accountType': contrato.tipocuentanomina or '',
        'payrollAccount': contrato.cuentanomina or '',
        'costCenter': contrato.idcosto.idcosto if contrato.idcosto else '',
        'subCostCenter': contrato.idsubcosto.idsubcosto if contrato.idsubcosto else '',
        'eps': contrato.codeps.identidad if contrato.codeps else '',
        'pensionFund': contrato.codafp.identidad if contrato.codafp else '',
        'CesanFund': contrato.codccf.identidad if contrato.codccf else '',
        'livingPlace': contrato.auxiliotransporte ,
        'arlWorkCenter': contrato.centrotrabajo.centrotrabajo if contrato.centrotrabajo else '',
        'workPlace': contrato.idsede.idsede if contrato.idsede else '',
        'contributor': contrato.tipocotizante.tipocotizante if contrato.tipocotizante else '',
        'subContributor': contrato.subtipocotizante.subtipocotizante if contrato.subtipocotizante else '',
    }


    # ✅ Limpieza de valores tipo "no data", "sin dato", "n/a", etc.
    data = {
        k: ("" if isinstance(v, str) and v.strip().lower() in ["no data", "sin dato", "n/a", "none", "ninguno"] else v)
        for k, v in data.items()
    }
    if request.method == 'POST':
        form = ReactivationForm(request.POST,idempresa=idempresa,empleado_actual = True , idcontrato = id)

        if form.is_valid():
            print('-------------')
            print(request.POST)
            print('-------------')
            response = HttpResponse()
            response['X-Up-Accept-Layer'] = 'true'  #Indica a Unpoly que acepte (cierre) el modal
            response['X-Up-icon'] = 'success'  # URL para recargar la página principal   
            response['X-Up-Message'] = 'La reactivación del empleado se realizó correctamente.' 
            response['X-Up-Location'] = reverse('companies:reactivation')           
            return response
        
        else: 
            for field, errors in form.errors.items():
                for error in errors:
                    response = HttpResponse()
                    response['X-Up-Accept-Layer'] = 'true'  #Indica a Unpoly que acepte (cierre) el modal
                    response['X-Up-icon'] = 'success'  # URL para recargar la página principal   
                    response['X-Up-message'] = f"Error en el campo '{field}': {error}"    
                    response['X-Up-Location'] = reverse('companies:reactivation')           
                    return response
                
            response = HttpResponse()
            response['X-Up-Accept-Layer'] = 'true'  #Indica a Unpoly que acepte (cierre) el modal
            response['X-Up-icon'] = 'success'  # URL para recargar la página principal   
            response['X-Up-message'] = "Todo lo que podía fallar, falló."    
            response['X-Up-Location'] = reverse('companies:reactivation')           
            return response
    else:
        form = ReactivationForm(idempresa=idempresa , initial=data , empleado_actual = True , idcontrato = id)
    return render(request, './companies/partials/reactivation_modal.html',{'form':form , 'idc':idc})




def estilo_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def ajustar_columnas(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width



@login_required
@role_required('company', 'accountant')
def reactivation_doc(request):
    # ------------------------
    # FUNCIONES AUXILIARES
    # ------------------------
    def clean_value(value):
        if isinstance(value, str) and value.strip().lower() in ["no data", "sin dato", "n/a", "none", "ninguno"]:
            return ""
        return value

    def crear_hoja(wb, title, headers, rows):
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        ajustar_columnas(ws)
        estilo_header(ws)
        return ws

    # ------------------------
    # DATOS DEL USUARIO
    # ------------------------
    usuario = request.session.get('usuario', {})
    idempresa = usuario.get('idempresa')

    # ------------------------
    # CONSULTAS DE EMPLEADOS
    # ------------------------
    contratos_activos = Contratos.objects.filter(
        idempleado=OuterRef('pk'),
        estadocontrato=1,
        id_empresa_id=idempresa
    )
    contratos_historicos = Contratos.objects.filter(
        idempleado=OuterRef('pk'),
        id_empresa_id=idempresa
    )
    ultimo_contrato = contratos_historicos.order_by('-idcontrato')

    empleados = (
        Contratosemp.objects.filter(id_empresa_id=idempresa)
        .annotate(
            tiene_contrato_activo=Exists(contratos_activos),
            tiene_contrato=Exists(contratos_historicos),
            salario=Subquery(ultimo_contrato.values('salario')[:1])
        )
        .filter(tiene_contrato_activo=False, tiene_contrato=True)
    )

    # Preparar datos de empleados
    empleados_data = []
    for e in empleados:
        nombre_completo = " ".join(filter(None, [e.pnombre, e.snombre, e.papellido, e.sapellido]))
        empleados_data.append([
            e.docidentidad,
            clean_value(nombre_completo.strip()),
            e.salario
        ])

    # ------------------------
    # CREAR EXCEL
    # ------------------------
    wb = Workbook()
    
    # Reusar la hoja inicial para "Plantilla"
    ws = wb.active
    ws.title = "Plantilla"

    plantilla_headers = [
        "Documento","Fecha de ingreso", "Fecha de retiro", "Salario",
        'Tipo Salario - ID', "Modalidad Salario - ID", "Cargo - ID", "Tipo de Contrato - ID",
        "Tipo de Nómina - ID", "Modelo de Contrato - ID", "Lugar de trabajo - ID", "Forma de pago - ID",
        "Banco de la Cuenta - ID", "Tipo de Cuenta - ID", "Cuenta de Nómina", "Eps - ID", "Pension - ID",
        "Fondo Cesantias - ID", "Centro de Trabajo ARL - ID", "Sede de Trabajo - ID", "Tipo de Cotizante - ID",
        "Subtipo de Cotizante - ID"
    ]
    ws.append(plantilla_headers)
    ajustar_columnas(ws)
    estilo_header(ws)

    # Hoja 2: Empleados
    crear_hoja(wb, "Empleados", ["Documento", "Nombre", "Salario anterior"], empleados_data)

    # Hoja 3: Cargos
    cargos = Cargos.objects.filter(id_empresa__idempresa=idempresa).exclude(idcargo=241).order_by('idcargo')
    crear_hoja(wb, "Cargos", ["ID", "Nombre"], [[c.idcargo, c.nombrecargo] for c in cargos])

    # Hoja 4: Entidades
    entidades = Entidadessegsocial.objects.exclude(
        codigo__in=['000', '9999', '9988', '9998']
    ).exclude(nit__in=['', None]).order_by('entidad')
    crear_hoja(wb, "Entidades", ["ID", "Codigo", "Nombre", "Tipo"], 
                [[e.identidad, e.codigo, e.entidad, e.tipoentidad] for e in entidades])

    # Hoja 5: Centros de Costos
    costos = Costos.objects.filter(id_empresa__idempresa=idempresa).exclude(grupocontable=0, suficosto=0).order_by('idcosto')
    crear_hoja(wb, "Centros de Costos", ["ID", "Nombre", "Grupo"], [[c.idcosto, c.nomcosto, c.grupocontable] for c in costos])

    # Hoja 6: Bancos
    costos = Bancos.objects.all().order_by('idbanco')
    crear_hoja(wb, "Bancos", ["ID", "Nombre", "Codigo"], [[c.idbanco, c.nombanco, c.codbanco] for c in costos])

    # Hoja de Choices
    choices_ws = wb.create_sheet(title="Choices")

    # Función auxiliar para escribir un bloque de choices
    def escribir_choices(ws, titulo, data, start_row):
        ws.cell(row=start_row, column=1, value=titulo)
        row = start_row + 1
        for key, value in data:
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        return row + 1  # Dejar una fila vacía entre bloques

    # Empezamos en la fila 1
    fila = 1
    fila = escribir_choices(choices_ws, "Modalidad Salario", ModalidadSalario, fila)
    fila = escribir_choices(choices_ws, "Forma de Pago", FormaPago, fila)
    fila = escribir_choices(choices_ws, "Tipo de Cuenta", TipoCcuenta, fila)
    
    # Ajustar columnas y estilo
    ajustar_columnas(choices_ws)
    estilo_header(choices_ws)
    # ------------------------
    # RESPUESTA HTTP
    # ------------------------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reactivacion_empleados.xlsx"'
    wb.save(response)
    return response




@login_required
@role_required('company', 'accountant')
def reactivation_data(request):


    return render(request, './companies/partials/reactivation_data.html')



