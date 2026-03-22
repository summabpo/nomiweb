from django.shortcuts import render, redirect, get_object_or_404
from apps.components.decorators import custom_login_required ,custom_permission
from django.db.models import OuterRef, Exists, Subquery
from apps.components.decorators import  role_required
from django.contrib.auth.decorators import login_required
from apps.companies.forms.ReactivationForm import ReactivationForm
from django.http import HttpResponse
from django.urls import reverse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment , PatternFill , Border, Side
import random
from django.utils.dateparse import parse_date

from apps.common.models import (
    # EMPLEADO
    Contratosemp,
    Tipodocumento,
    Ciudades,
    Paises,
    Profesiones,

    # CONTRATO
    Contratos,
    Cargos,
    Tipocontrato,
    Tipodenomina,
    Bancos,
    Centrotrabajo,
    Tiposdecotizantes,
    Subtipocotizantes,
    Tiposalario,
    ModelosContratos,
    Sedes,

    # SEGURIDAD SOCIAL
    Entidadessegsocial,

    # COSTOS (por si los usas después)
    Costos,
    Subcostos,

    # EMPRESA (clave)
    Empresa
)


ModalidadSalario = (
    ('1', 'Variable'),
    ('2', 'Fijo'),
    ('3', 'Mixto'),
)

FormaPago = (
    ('1', 'Abono a cuenta'),
    ('2', 'Cheque'),
    ('3', 'Efectivo'),
    ('4', 'Transferencia electrónica'),
)


TipoCcuenta = [
    ('ahorros', 'Ahorros'),
    ('corriente', 'Corriente'),
]

Tipo_sangre= [
    ('', '-----'), 
    ('OP', 'O +'), 
    ('ON', 'O -'), 
    ('AN', 'A -'), 
    ('AP', 'A +'), 
    ('BP', 'B +'), 
    ('BN', 'B -'), 
    ('ABP', 'AB +'), 
    ('ABN', 'AB -')
    ]


Educacion = [
            ('primaria', 'Primaria'), 
            ('Bachiller', 'Bachiller'), 
            ('bachillerinc', 'Bachiller Incompleto'), 
            ('tecnico', 'Técnico'), 
            ('tecnologo', 'Tecnólogo'), 
            ('universitario', 'Universitario'), 
            ('universitarioinc', 'Universitario Incompleto'), 
            ('postgrado', 'Postgrado'), 
            ('magister', 'Magíster')
        ]

estado=[
    ('soltero', 'Soltero'), 
    ('casado', 'Casado'), 
    ('viudo', 'Viudo'), 
    ('divorciado', 'Divorciado'), 
    ('unionlibre', 'Unión Libre')
    ]


tipo_documento=[
        ('1','Registro Civil'),
        ('2','Tarjeta de identidad'),
        ('3','Cedula de ciudadanía'),
        ('4','Tarjeta de extranjería'),
        ('5','Cédula de extranjería'),
        ('6','NIT'),
        ('7','Pasaporte'),
        ('8','Documento de indentificación extranjero'),
        ('9','Permiso especial de permanencia'),
        ('10','NIT otro país'),
        ('11','NUIP'),
    ]


Tiponomina = [
    ('1','Mensual'),
    ('2','Quincenal'),
    ('3','Por Horas'),
    ('4','Primas'),
    ('5','Cesantias'),
    ('6','Adicional'),
    ('7','Vacaciones'),
    ('8','Liquidación'),
    ('9','Catorcenal'),
    ('10','Int. de Cesantias'),
    ('11','Semanal'),
    ('12','Provisiones'),
]

tipo_salario = [
    ('1','Normal'),
    ('2','Integral'),
    ('3','Horas'),
    ('4','Variable'),
]

tipo_cotizante = [ 
    ('ND','No Data'),
    ('23','Estudiantes aportes solo riesgos laborales'),
    ('01','Dependiente'),
    ('02','Servicio doméstico'),
    ('03','Independiente'),
    ('04','Madre comunitaria'),
    ('12','Aprendices SENA etapa lectiva'),
    ('15','Desempleado con subsidio CCF'),
    ('16','Independiente agremiado o asociado'),
    ('18','Funcionarios públicos sin tope máximo IBC'),
    ('19','Aprendices SENA etapa productiva'),
    ('20','Estudiantes (Régimen Especial Ley 789/2002)'),
    ('21','Estudiantes de postgrado en salud (Decreto 190/1996)'),
    ('22','Profesor establecimiento particular'),
    ('30','Dependiente entidades públicas régimen especial salud'),
    ('31','Cooperado o PreCooperativa de trabajo asociado'),
    ('32','Miembro carrera diplomática/consular extranjero'),
    ('33','Beneficiario Fondo Solidaridad Pensional'),
    ('34','Concejal/edil Bogotá amparado por póliza salud'),
    ('35','Concejal municipal o distrital'),
    ('36','Edil Junta Administradora Local'),
    ('40','Beneficiario UPC adicional'),
    ('41','Independiente sin ingresos pago por terceros'),
    ('42','Pago solo salud (Art. 2, Ley 1250/2008)'),
    ('43','Independiente no obligado pensión pago terceros'),
    ('44','Dependiente Empleo Emergencia ≥ 1 mes'),
    ('45','Dependiente Empleo Emergencia < 1 mes'),
    ('47','Trabajador entidad Sistema Gral. Participaciones'),
    ('51','Trabajador tiempo parcial (nuevo 2025)'),
    ('57','Independiente voluntario ARL'),
    ('59','Contratista independiente contrato > 1 mes'),
    ('60','Edil JAL no beneficiario Fondo Solidaridad'),
    ('64','Trabajador penitenciario'),
    ('67','Voluntario Primera Respuesta solo ARL'),
    ('70','Promotor Servicio Social para la Paz'),
    ('72','Mujer aporte pensión por tercero (nuevo 2025)'),

]




def estilo_header(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def ajustar_columnas(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

def color_aleatorio_suave():
    r = lambda: random.randint(120, 200)  # base más visible
    return f"{r():02X}{r():02X}{r():02X}"


def aclarar_color(hex_color, factor=0.5):
    """
    factor: 0 = igual, 1 = blanco total
    """
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)

    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)

    return f"{r:02X}{g:02X}{b:02X}"

borde_negro = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)


@login_required
@role_required('company', 'accountant')
def masive_contract_modal(request):


    return render(request, './companies/partials/masive_contract_modal.html')



@login_required
@role_required('company', 'accountant')
def masive_contract_doc(request):
    # ------------------------
    # FUNCIONES AUXILIARES
    # ------------------------
    def clean_value(value):
        if isinstance(value, str) and value.strip().lower() in ["no data", "sin dato", "n/a", "none", "ninguno"]:
            return ""
        return value

    def crear_hoja(wb, title, headers, rows):
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for row in rows:
            ws.append(row)
        ajustar_columnas(ws)
        estilo_header(ws)
        return ws

    # ------------------------
    # DATOS DEL USUARIO
    # ------------------------
    usuario = request.session.get('usuario', {})
    idempresa = usuario.get('idempresa')

    # ------------------------
    # CONSULTAS DE EMPLEADOS
    # ------------------------
    contratos_activos = Contratos.objects.filter(
        idempleado=OuterRef('pk'),
        estadocontrato=1,
        id_empresa_id=idempresa
    )
    contratos_historicos = Contratos.objects.filter(
        idempleado=OuterRef('pk'),
        id_empresa_id=idempresa
    )
    ultimo_contrato = contratos_historicos.order_by('-idcontrato')

    empleados = (
        Contratosemp.objects.filter(id_empresa_id=idempresa)
        .annotate(
            tiene_contrato_activo=Exists(contratos_activos),
            tiene_contrato=Exists(contratos_historicos),
            salario=Subquery(ultimo_contrato.values('salario')[:1])
        )
        .filter(tiene_contrato_activo=False, tiene_contrato=True)
    )

    # Preparar datos de empleados
    empleados_data = []
    for e in empleados:
        nombre_completo = " ".join(filter(None, [e.pnombre, e.snombre, e.papellido, e.sapellido]))
        empleados_data.append([
            e.docidentidad,
            clean_value(nombre_completo.strip()),
            e.salario
        ])

    # ------------------------
    # CREAR EXCEL
    # ------------------------
    wb = Workbook()
    
    # Reusar la hoja inicial para "Plantilla"
    ws = wb.active
    ws.title = "Contratos"
    # Hoja 2: Empleados
    plantilla_headers = [
        "Documento","Fecha de ingreso", "Fecha de retiro", "Salario",
        'Tipo Salario - ID', "Modalidad Salario - ID", "Cargo - ID", "Tipo de Contrato - ID",
        "Tipo de Nómina - ID", "Modelo de Contrato - ID", "Lugar de trabajo - ID", "Forma de pago - ID",
        "Banco de la Cuenta - ID", "Tipo de Cuenta - ID", "Cuenta de Nómina", "Eps - ID", "Pension - ID",
        "Fondo Cesantias - ID", "Centro de Trabajo ARL - ID", "Sede de Trabajo - ID", "Tipo de Cotizante - ID",
        "Subtipo de Cotizante - ID"
    ]
    ws.append(plantilla_headers)
    ajustar_columnas(ws)
    estilo_header(ws)


    plantilla_headers = [
        "Documento","Tipo Documento", "Fecha de expedición", "Ciudad de expedición - ID",
        'Primer Nombre', "Segundo Nombre", "Primer Apellido", "Segundo Apellido",
        "Sexo", "Estatura (Mts)", "Estado Civil", "Peso (Kg)",
        "Fecha de Nacimiento", "Nivel Educativo", "Ciudad de Nacimiento", "Estrato", "País de Nacimiento - ID",
        "Libreta Militar", "Grupo Sanguíneo - ID", "Profesión - ID", "Dirección de Residencia",
        "E-mail","Ciudad de Residencia - ID","País de residencia - ID","Celular","Teléfono del Empleado",
        "Nombre de contacto" , "Celular de contacto" , "Tipo de relación",
        "Talla Pantalón" , "Talla Camisa" , "Talla Zapatos"
    ]

    # Hoja 2: Empleados
    crear_hoja(wb, "Hojas de vida", plantilla_headers,[])

    # Hoja 3: Cargos
    cargos = Cargos.objects.filter(id_empresa__idempresa=idempresa).exclude(idcargo=241).order_by('idcargo')
    crear_hoja(wb, "Cargos", ["ID", "Nombre"], [[c.idcargo, c.nombrecargo] for c in cargos])

    # Hoja 4: Entidades
    entidades = Entidadessegsocial.objects.exclude(
        codigo__in=['000', '9999', '9988', '9998']
    ).exclude(nit__in=['', None]).order_by('entidad')
    crear_hoja(wb, "Entidades", ["ID", "Codigo", "Nombre", "Tipo"], 
                [[e.identidad, e.codigo, e.entidad, e.tipoentidad] for e in entidades])

    # Hoja 5: Centros de Costos
    costos = Costos.objects.filter(id_empresa__idempresa=idempresa).exclude(grupocontable=0, suficosto=0).order_by('idcosto')
    crear_hoja(wb, "Centros de Costos", ["ID", "Nombre", "Grupo"], [[c.idcosto, c.nomcosto, c.grupocontable] for c in costos])

    # Hoja 6: Bancos
    costos = Bancos.objects.all().order_by('idbanco')
    crear_hoja(wb, "Bancos", ["ID", "Nombre", "Codigo"], [[c.idbanco, c.nombanco, c.codbanco] for c in costos])

    # Hoja 7: Ciudades
    ciudades = Ciudades.objects.all().order_by('idciudad')
    crear_hoja(wb, "Ciudades", ["ID", "Nombre", "Codigo","Departamento"], [[c.idciudad, c.ciudad,c.codciudad, c.departamento] for c in ciudades])


    # Hoja 8: 
    profesiones = Profesiones.objects.all().order_by('idprofesion')
    crear_hoja(wb, "Profesiones", ["ID", "Nombre"], [[p.idprofesion, p.profesion] for p in profesiones])


    # Hoja : Choices
    choices_ws = wb.create_sheet(title="Choices")

    # Función auxiliar para escribir un bloque de choices
    def escribir_choices(ws, titulo, data, start_row):

        def color_aleatorio_suave():
            r = lambda: random.randint(120, 200)
            return f"{r():02X}{r():02X}{r():02X}"

        def aclarar_color(hex_color, factor=0.7):
            r = int(hex_color[0:2], 16)
            g = int(hex_color[2:4], 16)
            b = int(hex_color[4:6], 16)

            r = int(r + (255 - r) * factor)
            g = int(g + (255 - g) * factor)
            b = int(b + (255 - b) * factor)

            return f"{r:02X}{g:02X}{b:02X}"

        # Borde negro
        borde = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # Colores
        color_base = color_aleatorio_suave()
        color_claro = aclarar_color(color_base)

        fill_titulo = PatternFill(start_color=color_base, end_color=color_base, fill_type="solid")
        fill_filas = PatternFill(start_color=color_claro, end_color=color_claro, fill_type="solid")

        # TÍTULO
        c1 = ws.cell(row=start_row, column=1, value=titulo)
        c2 = ws.cell(row=start_row, column=2)

        for c in (c1, c2):
            c.fill = fill_titulo
            c.font = Font(bold=True)
            c.border = borde

        row = start_row + 1

        # Normalizar data
        if isinstance(data, tuple) and len(data) == 1:
            data = data[0]

        # FILAS
        for key, value in data:
            c1 = ws.cell(row=row, column=1, value=key)
            c2 = ws.cell(row=row, column=2, value=value)

            for c in (c1, c2):
                c.fill = fill_filas
                c.border = borde

            row += 1

        return row + 1



    # Empezamos en la fila 1
    fila = 1
    fila = escribir_choices(choices_ws, "Tipo de documento de identidad", tipo_documento, fila)
    fila = escribir_choices(choices_ws, "Modalidad Salario", ModalidadSalario, fila)
    fila = escribir_choices(choices_ws, "Forma de Pago", FormaPago, fila)
    fila = escribir_choices(choices_ws, "Tipo de Cuenta", TipoCcuenta, fila)
    fila = escribir_choices(choices_ws, "Tipo sangre", Tipo_sangre, fila)
    fila = escribir_choices(choices_ws, "Nivel Educativo", Educacion, fila)
    fila = escribir_choices(choices_ws, "Estado Civil", estado, fila)
    fila = escribir_choices(choices_ws, "Tipo de Nómina", Tiponomina, fila)
    fila = escribir_choices(choices_ws, "Tipo Salario", tipo_salario, fila)
    fila = escribir_choices(choices_ws, "Tipo de Cotizante", tipo_cotizante, fila)

    # Ajustar columnas y estilo
    ajustar_columnas(choices_ws)
    estilo_header(choices_ws)
    # ------------------------
    # RESPUESTA HTTP
    # ------------------------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="reactivacion_empleados.xlsx"'
    wb.save(response)
    return response


def clean_value(value):
    if isinstance(value, str) and value.strip().lower() in ["no data", "sin dato", "n/a", "none", "ninguno"]:
        return None
    return value



def create_employee(row, empresa):
    errors = []
    
    try:
        # ------------------------
        # LIMPIEZA
        # ------------------------
        doc = clean_value(row.get("Documento"))

        if not doc:
            return {
                "status": "error",
                "documento": None,
                "errors": ["Documento es obligatorio"]
            }

        # ------------------------
        # VALIDACIÓN DUPLICADO
        # ------------------------
        existing = Contratosemp.objects.filter(
            docidentidad=doc,
            id_empresa=empresa
        ).exists()

        if existing:
            return {
                "status": "duplicate",
                "documento": doc,
                "errors": ["Empleado ya existe en la empresa"]
            }

        # ------------------------
        # FOREIGN KEYS
        # ------------------------
        tipo_doc_id = row.get("Tipo Documento")
        tipo_doc = Tipodocumento.objects.filter(id=tipo_doc_id).first()
        if not tipo_doc:
            errors.append(f"Tipo Documento inválido: {tipo_doc_id}")

        ciudad_exp_id = row.get("Ciudad de expedición - ID")
        ciudad_exp = Ciudades.objects.filter(id=ciudad_exp_id).first()
        if not ciudad_exp:
            errors.append(f"Ciudad expedición inválida: {ciudad_exp_id}")

        ciudad_nac_id = row.get("Ciudad de Nacimiento")
        ciudad_nac = Ciudades.objects.filter(id=ciudad_nac_id).first()
        if not ciudad_nac:
            errors.append(f"Ciudad nacimiento inválida: {ciudad_nac_id}")

        ciudad_res_id = row.get("Ciudad de Residencia - ID")
        ciudad_res = Ciudades.objects.filter(id=ciudad_res_id).first()
        if not ciudad_res:
            errors.append(f"Ciudad residencia inválida: {ciudad_res_id}")

        pais_nac_id = row.get("País de Nacimiento - ID")
        pais_nac = Paises.objects.filter(id=pais_nac_id).first()
        if not pais_nac:
            errors.append(f"País nacimiento inválido: {pais_nac_id}")

        pais_res_id = row.get("País de residencia - ID")
        pais_res = Paises.objects.filter(id=pais_res_id).first()
        if not pais_res:
            errors.append(f"País residencia inválido: {pais_res_id}")

        # ------------------------
        # FECHAS
        # ------------------------
        fecha_nac = parse_date(str(row.get("Fecha de Nacimiento")))
        if not fecha_nac:
            errors.append("Fecha de nacimiento inválida")

        fecha_exp = parse_date(str(row.get("Fecha de expedición")))
        if not fecha_exp:
            errors.append("Fecha de expedición inválida")

        # ------------------------
        # CAMPOS OBLIGATORIOS
        # ------------------------
        pnombre = clean_value(row.get("Primer Nombre"))
        papellido = clean_value(row.get("Primer Apellido"))

        if not pnombre:
            errors.append("Primer Nombre es obligatorio")

        if not papellido:
            errors.append("Primer Apellido es obligatorio")

        # ------------------------
        # SI HAY ERRORES → NO CREA
        # ------------------------
        if errors:
            return {
                "status": "error",
                "documento": doc,
                "errors": errors
            }

        # ------------------------
        # CREACIÓN
        # ------------------------
        Contratosemp.objects.create(
            docidentidad=doc,
            tipodocident=tipo_doc,

            pnombre=pnombre,
            snombre=clean_value(row.get("Segundo Nombre")),
            papellido=papellido,
            sapellido=clean_value(row.get("Segundo Apellido")),

            email=clean_value(row.get("E-mail")),
            telefonoempleado=clean_value(row.get("Teléfono del Empleado")),
            celular=clean_value(row.get("Celular")),
            direccionempleado=clean_value(row.get("Dirección de Residencia")),

            sexo=clean_value(row.get("Sexo")),
            fechanac=fecha_nac,

            ciudadnacimiento=ciudad_nac,
            paisnacimiento=pais_nac,
            ciudadresidencia=ciudad_res,
            paisresidencia=pais_res,

            estadocivil=clean_value(row.get("Estado Civil")),

            profesion=clean_value(row.get("Profesión - ID")),
            niveleducativo=clean_value(row.get("Nivel Educativo")),

            estatura=clean_value(row.get("Estatura (Mts)")),
            peso=clean_value(row.get("Peso (Kg)")),
            gruposanguineo=clean_value(row.get("Grupo Sanguíneo - ID")),

            fechaexpedicion=fecha_exp,
            ciudadexpedicion=ciudad_exp,

            dotpantalon=clean_value(row.get("Talla Pantalón")),
            dotcamisa=clean_value(row.get("Talla Camisa")),
            dotzapatos=clean_value(row.get("Talla Zapatos")),

            estrato=clean_value(row.get("Estrato")),
            numlibretamil=clean_value(row.get("Libreta Militar")),

            contact_name=clean_value(row.get("Nombre de contacto")),
            contact_cell_phone=clean_value(row.get("Celular de contacto")),
            contact_relationship=clean_value(row.get("Tipo de relación")),

            id_empresa=empresa
        )

        return {
            "status": "created",
            "documento": doc,
            "errors": []
        }

    except Exception as e:
        return {
            "status": "exception",
            "documento": row.get("Documento"),
            "errors": [str(e)]
        }




def create_contract(row, empresa):
    errors = []

    try:
        # ------------------------
        # DOCUMENTO (CLAVE)
        # ------------------------
        doc = clean_value(row.get("Documento"))

        if not doc:
            return {
                "status": "error",
                "documento": None,
                "errors": ["Documento es obligatorio"]
            }

        # ------------------------
        # EMPLEADO EXISTENTE
        # ------------------------
        empleado = Contratosemp.objects.filter(
            docidentidad=doc,
            id_empresa=empresa
        ).first()

        if not empleado:
            return {
                "status": "error",
                "documento": doc,
                "errors": ["Empleado no existe en la empresa"]
            }

        # ------------------------
        # VALIDACIÓN DUPLICADO (opcional)
        # ------------------------
        existing = Contratos.objects.filter(
            idempleado=empleado,
            fechainiciocontrato=row.get("Fecha de ingreso")
        ).exists()

        if existing:
            return {
                "status": "duplicate",
                "documento": doc,
                "errors": ["Contrato ya existe para este empleado en esa fecha"]
            }

        # ------------------------
        # FOREIGN KEYS
        # ------------------------
        cargo = Cargos.objects.filter(id=row.get("Cargo - ID")).first()
        if not cargo:
            errors.append(f"Cargo inválido: {row.get('Cargo - ID')}")

        tipo_contrato = Tipocontrato.objects.filter(id=row.get("Tipo de Contrato - ID")).first()
        if not tipo_contrato:
            errors.append("Tipo de contrato inválido")

        tipo_nomina = Tipodenomina.objects.filter(id=row.get("Tipo de Nómina - ID")).first()
        if not tipo_nomina:
            errors.append("Tipo de nómina inválido")

        banco = Bancos.objects.filter(id=row.get("Banco de la Cuenta - ID")).first()

        centro = Centrotrabajo.objects.filter(id=row.get("Centro de Trabajo ARL - ID")).first()
        if not centro:
            errors.append("Centro de trabajo inválido")

        sede = Sedes.objects.filter(id=row.get("Sede de Trabajo - ID")).first()

        ciudad = Ciudades.objects.filter(id=row.get("Lugar de trabajo - ID")).first()
        if not ciudad:
            errors.append("Lugar de trabajo inválido")

        tipo_cotizante = Tiposdecotizantes.objects.filter(id=row.get("Tipo de Cotizante - ID")).first()
        if not tipo_cotizante:
            errors.append("Tipo de cotizante inválido")

        subtipo_cotizante = Subtipocotizantes.objects.filter(id=row.get("Subtipo de Cotizante - ID")).first()
        if not subtipo_cotizante:
            errors.append("Subtipo de cotizante inválido")

        tipo_salario = Tiposalario.objects.filter(id=row.get("Tipo Salario - ID")).first()

        modelo = ModelosContratos.objects.filter(id=row.get("Modelo de Contrato - ID")).first()
        if not modelo:
            errors.append("Modelo de contrato inválido")

        eps = Entidadessegsocial.objects.filter(id=row.get("Eps - ID")).first()
        if not eps:
            errors.append("EPS inválida")

        pension = Entidadessegsocial.objects.filter(id=row.get("Pension - ID")).first()
        cesantias = Entidadessegsocial.objects.filter(id=row.get("Fondo Cesantias - ID")).first()

        # ------------------------
        # FECHAS
        # ------------------------
        fecha_inicio = parse_date(str(row.get("Fecha de ingreso")))
        if not fecha_inicio:
            errors.append("Fecha de ingreso inválida")

        fecha_fin = parse_date(str(row.get("Fecha de retiro"))) if row.get("Fecha de retiro") else None

        # ------------------------
        # SALARIO
        # ------------------------
        salario = clean_value(row.get("Salario"))
        if not salario:
            errors.append("Salario es obligatorio")

        # ------------------------
        # SI HAY ERRORES
        # ------------------------
        if errors:
            return {
                "status": "error",
                "documento": doc,
                "errors": errors
            }

        # ------------------------
        # CREACIÓN
        # ------------------------
        Contratos.objects.create(
            idempleado=empleado,
            cargo=cargo,
            fechainiciocontrato=fecha_inicio,
            fechafincontrato=fecha_fin,
            tipocontrato=tipo_contrato,
            tiponomina=tipo_nomina,
            bancocuenta=banco,
            cuentanomina=clean_value(row.get("Cuenta de Nómina")),
            tipocuentanomina=clean_value(row.get("Tipo de Cuenta - ID")),
            centrotrabajo=centro,
            ciudadcontratacion=ciudad,
            salario=salario,
            tipocotizante=tipo_cotizante,
            subtipocotizante=subtipo_cotizante,
            formapago=clean_value(row.get("Forma de pago - ID")),
            tiposalario=tipo_salario,
            idsede=sede,
            codeps=eps,
            codafp=pension,
            fondocesantias=cesantias,
            idmodelo=modelo,
            estadocontrato=1,  # activo por defecto
            id_empresa=empresa
        )

        return {
            "status": "created",
            "documento": doc,
            "errors": []
        }

    except Exception as e:
        return {
            "status": "exception",
            "documento": row.get("Documento"),
            "errors": [str(e)]
        }