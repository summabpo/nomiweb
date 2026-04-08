from django.shortcuts import render
from apps.components.decorators import  role_required
from django.contrib.auth.decorators import login_required
from apps.companies.forms.updatesalaryForm import updatesalaryForm
from django.http import JsonResponse , HttpResponse
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.dimensions import DimensionHolder
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.http import JsonResponse
from apps.common.models import NovSalarios, Contratos ,HistorialSalario
from decimal import Decimal

from django.db import transaction

error_messages = {
    '1': "Id contrato inválido o faltante",
    '3': "Contrato no encontrado",
    '4': "El contrato no está activo",
    '6': "El contrato no pertenece a la empresa actual",
    '7': "Fila vacía o incompleta",
    '8': "El nuevo salario debe ser mayor que cero",
    '13': "El contrato está cerrado o inactivo",
    '14': "El nuevo salario no puede ser negativo",
    '19': "El nuevo salario es igual al salario actual",
    '20': "Ya existe un ajuste salarial con la misma fecha",
    '21': "Ajuste salarial duplicado en el archivo",
    'general': "Error general: Error procesando el archivo"
}


data = [
    79666771, 79852078, 1031136265, 79811576, 80008103, 1024569199, 1010208848,
    80027767, 79701956, 1000987982, 1069481010, 1024599316, 7335056, 52061581,
    1069053766, 79838923, 79770362, 79218724, 1030611165, 52534684, 74795301,
    7141946, 80374258, 1026583561, 1033821192, 1030566847, 79424564, 1000128752,
    79611326, 80813459, 1013590759, 1012386632, 1023947902, 1024489806, 1012375148,
    80829168, 76044275, 80773287, 79467287, 1024481445, 80156553, 1025523195,
    1033695097, 7279172, 1102816899, 1076624253, 1024539316, 79977684, 1030644442,
    53167380, 79920433, 1110538988, 63437190, 1016042327, 79701697, 1033800409,
    1024475880, 1010028274, 1023888780, 80919740, 79977830, 1023939049, 93299546,
    1022397439, 1026551048, 52243797, 79495213, 88267948, 1032419900, 1010180190,
    80770418, 79433053, 80913621, 16803629, 1000128092, 1012399610, 79855526,
    1023929073, 1032412194, 79709773, 1071171615, 1193155768, 1022945507, 80069702,
    80215910, 80212300, 79993029, 1143446588, 80219635, 1012341454, 79739546,
    1010141997, 79816324, 79048766, 80902857, 80155677, 80114515, 1033729671,
    1010246967, 1031127024, 79409833, 79985460, 1023873938, 1016023985, 80793172,
    79658305, 79629030, 80768640, 80112624, 1022941442, 1003026446, 79836196,
    1024485103, 79871049, 80758413, 1015435501, 1014221145, 1069716680, 1000618295,
    79629649, 1024505569, 53044370, 52902133, 1023891206, 1030700708, 52279321,
    80097898, 1012324574, 1024517286, 52730433, 1012346945, 1012354573, 1076220721,
    1077969333, 52318550, 80130541, 1026561070, 1024587545, 1032415379, 1013634848,
    80827073, 79430753, 79708074, 80100733, 80881371, 78741051, 1015430699,
    1023955542, 80274418, 80210764, 79801213, 1000126840, 1015400696, 1022941380,
    1073713507, 1075674385, 1022385525, 1078369310, 1022997377, 1109244475,
    1024470002, 1000719629, 79456369, 80769118, 1030572353, 1033757273, 1010221013,
    79772179, 1016029972, 79819146, 1010963178, 1010245056, 1034656439, 1026290929,
    80140545, 1033729825, 1073687466, 1024497450, 80149199, 79726920, 1013601903,
    1022330906, 80812020, 79988126, 1001117175, 1024595059, 1028861336, 1014191903,
    1030520433, 1012426990, 11309970, 1108934346, 1074133128, 1070625452, 11228959,
    1071549318, 1074134723, 1072494090, 1069755714, 1007570516, 1072961638,
    1072196507, 1072423919, 1072423301, 1070615274, 79065837, 1072495200,
    1003652869, 1007184466, 81741405, 11220405, 1070620646, 3197384, 1069762417,
    1069742841, 1075626898, 1075629124, 1070626593, 1072431321, 1003530318,
    1003530319, 1070329946, 11315234, 3197439, 1070324887, 79065389, 1072426348,
    1074131611, 1074131716, 1070327299, 1070600963, 1070331340, 81741039,
    1022366209, 1070622462, 1075629342, 11413845, 79064342, 1069737042, 11206413,
    11223495, 1007316757, 1000697007, 1026306019, 1069736457, 1072496426,
    1099548343, 1070325660, 1070324106, 3170953, 1069282975, 3171605, 79761580,
    1074132807, 1069712305, 1233501202, 1074136531, 1069746170, 11259763,
    1069750382, 9801619, 1072423313, 1070331502, 1069750872, 1074135310,
    1071550622, 79838616, 1069749029, 1070611024, 1007847085, 1069720783,
    1070324421, 1078826939, 1004155776, 1039623846, 1068973265, 1016028457,
    1069756956
]

@login_required
@role_required('company','accountant')
def update_salary(request):
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']

    # Obtener fecha actual
    from datetime import date
    
    fecha_hoy = date.today()
    
    novsalarios = (
        NovSalarios.objects
        .select_related(
            'idcontrato',
            'idcontrato__idempleado'
        )
        .only(
            'idcambiosalario',
            'salarioactual',
            'nuevosalario',
            'fechanuevosalario',
            'idcontrato__idcontrato',
            'idcontrato__idempleado__docidentidad',
            'idcontrato__idempleado__pnombre',
            'idcontrato__idempleado__papellido',
        )
        .filter(
            idcontrato__id_empresa=idempresa
        )
        .order_by('idcambiosalario')
    )
    

    # for doc in data:
    #     obj = novsalarios.filter(
    #         idcontrato__idempleado__docidentidad=doc,
    #         fechanuevosalario=date(2026, 1, 1)
    #     ).first()
        
    #     if obj:
    #         obj.fechanuevosalario = date(2026, 1, 1)
    #         obj.save()
    #     else:
    #         #print(f"✘ {doc} -> NO existe")
    #         cc = Contratos.objects.filter(
    #             idempleado__docidentidad=doc
    #         ).order_by('-idcontrato').first()  # o '-id' según tu PK

    #         if cc:
    #             NovSalarios.objects.create(
    #                 idcontrato=cc,
    #                 salarioactual=1506052,
    #                 nuevosalario=1750905,
    #                 fechanuevosalario=date(2026, 1, 1),
    #                 tiposalario=1
    #             )
    #         else:
    #             print(f"No se encontró contrato para {doc}")




    for i in novsalarios:
        contrato = Contratos.objects.get(idcontrato=i.idcontrato.idcontrato)
        
        if i.fechanuevosalario <= fecha_hoy and contrato.salario != i.nuevosalario:
            contrato.salario = i.nuevosalario
            contrato.save()
    
    return render(request, './companies/update_salary.html', {'novsalarios': novsalarios})




def actualizacion():


    return('ok')


@login_required
@role_required('company','accountant')
def update_salary_add(request):
    existe_historial = False
    usuario = request.session.get('usuario', {})
    idempresa = usuario['idempresa']

    form = updatesalaryForm(idempresa = idempresa)
    if request.method == 'POST':
        
        form = updatesalaryForm(request.POST , idempresa=idempresa )
        if form.is_valid():
            
            print(request.POST)

            contrato_id = form.cleaned_data['contract']
            nuevosalario = form.cleaned_data['Salario_nuevo']
            fecha_nuevo = form.cleaned_data['fecha_nuevo']
            salario_actual = form.cleaned_data['Salario_Actual']
            salario_nuevo = form.cleaned_data['Salario_nuevo']
            fecha_nuevo = form.cleaned_data['fecha_nuevo']
            tiposalario = form.cleaned_data['contractType']

            contrato = Contratos.objects.get(idcontrato = form.cleaned_data['contract'] )
            

            if nuevosalario > contrato.salario : 

                dataaux =  NovSalarios.objects.filter(nuevosalario = contrato.salario , idcontrato_id = form.cleaned_data['contract'] ).first()
                
                # validar existencia
                if dataaux : 
                    existe_historial = HistorialSalario.objects.filter(
                        contrato_id=contrato_id,
                        salario=salario_nuevo,
                        fecha_inicio=dataaux.fechanuevosalario,
                        fecha_fin=fecha_nuevo
                    ).exists()
                


                existe_novsalario = NovSalarios.objects.filter(
                    idcontrato_id=contrato_id,
                    salarioactual=salario_actual,
                    nuevosalario=salario_nuevo,
                    fechanuevosalario=fecha_nuevo,
                    tiposalario=tiposalario,
                ).exists()

                if not existe_novsalario:

                    with transaction.atomic():

                        novsalario = NovSalarios.objects.create(
                            idcontrato_id=contrato_id,
                            salarioactual=salario_actual,
                            nuevosalario=salario_nuevo,
                            fechanuevosalario=fecha_nuevo,
                            tiposalario=tiposalario,
                        )

                        if  not existe_historial  : 
                            historial = HistorialSalario.objects.create(
                                contrato_id=contrato_id,
                                salario=contrato.salario,
                                fecha_inicio=novsalario.fechanuevosalario,
                                es_actual=True,
                            )
                        else:
                            historial = HistorialSalario.objects.create(
                                contrato_id=contrato_id,
                                salario=contrato.salario,
                                fecha_inicio=dataaux.fechanuevosalario,
                                fecha_fin=fecha_nuevo,
                                es_actual=True,
                            )
                            

                        response = HttpResponse()
                        response['X-Up-Accept-Layer'] = 'true'  #Indica a Unpoly que acepte (cierre) el modal
                        response['X-Up-icon'] = 'success'  # URL para recargar la página principal   
                        response['X-Up-message'] = 'Nuevo salario guardada exitosamente'    
                        response['X-Up-Location'] = reverse('companies:update_salary')           
                        return response
                else:
                    
                    response = HttpResponse()
                    response['X-Up-Accept-Layer'] = 'true'  #Indica a Unpoly que acepte (cierre) el modal
                    response['X-Up-icon'] = 'warning'  # URL para recargar la página principal   
                    response['X-Up-message'] = 'ups , parece que esa data ya la tenemos'    
                    response['X-Up-Location'] = reverse('companies:update_salary')           
                    return response
        else:
            # En caso de que el formulario no sea válido, mostrar los errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    print(request, f"Error en {field}: {error}")   
    return render (request, './companies/partials/update_salary_add.html',{'form':form})


@login_required
@role_required('company', 'accountant')
def update_salary_add_masive(request):

    usuario = request.session.get('usuario', {})
    idempresa = int(usuario['idempresa'])

    errores = []
    pending_rows = []   # datos válidos en memoria
    has_errors = False  # flag global

    if request.method == 'POST':
        file = request.FILES.get("file")
        if not file:
            return JsonResponse({"errors": ["Archivo no proporcionado"]}, status=400)

        # Leer archivo
        try:

            file.seek(0)

            if file.name.endswith('.csv'):
                df = pd.read_csv(file)

            elif file.name.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file, engine='openpyxl')

            else:

                return JsonResponse(
                    {"errors": ["Formato no soportado"]},
                    status=400
                )

        except Exception as e:
            print("ERROR REAL:", e)
            return JsonResponse(
                {"errors": [str(e)]},
                status=400
            )
        
        cargados = set()  # detectar duplicados en archivo

        # =========================
        # 1️VALIDAR TODO
        # =========================
        for index, row in df.iterrows():
            row_errors = []  # se reinicia por fila
            try:
                id_contrato = int(row.get('ID Contrato'))
                nuevo_salario = float(row.get('Nuevo Salario'))
                fecha = row.get('Fecha')
            except Exception:
                row_errors.append('7')

            contrato = None
            if not row_errors:
                contrato = Contratos.objects.filter(
                    idcontrato=id_contrato,
                    id_empresa_id=idempresa
                ).first()

                if not contrato:
                    row_errors.append('3')
                elif contrato.estadocontrato != 1:
                    row_errors.append('4')

                if nuevo_salario <= 0:
                    row_errors.append('8')

                if contrato and nuevo_salario == float(contrato.salario):
                    row_errors.append('19')

                if contrato and NovSalarios.objects.filter(
                    idcontrato=contrato,
                    fechanuevosalario=fecha
                ).exists():
                    row_errors.append('20')

                key = (id_contrato, fecha)
                if key in cargados:
                    row_errors.append('21')
                else:
                    cargados.add(key)

            # Si hay errores → marcar flag global
            if row_errors:
                has_errors = True
                for err in row_errors:
                    print('------------------')
                    print(error_messages[err])

                    errores.append({
                        "fila": index + 1,
                        "id_contrato": id_contrato,
                        "error": error_messages[err]
                    })
            else:
                pending_rows.append({
                    "contrato": contrato,
                    "nuevo_salario": nuevo_salario,
                    "fecha": fecha
                })

        # Si hubo cualquier error → NO guardar
        if has_errors:
                        
            return render(
                request,
                './companies/partials/update_salary_add_masive_error.html',
                {"errores": errores,
                "message": "No se guardó ningún registro"}
            )

        # =========================
        # 2GUARDAR (SIN ERRORES)
        # =========================
        # =========================
        # GUARDAR
        # =========================
        try:
            with transaction.atomic():
                for r in pending_rows:
                    NovSalarios.objects.create(
                        idcontrato=r["contrato"],
                        salarioactual=r["contrato"].salario,
                        nuevosalario=r["nuevo_salario"],
                        fechanuevosalario=r["fecha"],
                        tiposalario=r["contrato"].tiposalario.idtiposalario
                    )

        except Exception as e:

            return render(
                request,
                './companies/partials/update_salary_add_masive_error.html',
                {
                    "errores": [{"error": str(e)}],
                    "message": "Ocurrió un error al guardar."
                }
            )

        # =========================
        # SUCCESS
        # =========================
        return render(
            request,
            './companies/partials/update_salary_add_masive_success.html',
            {
                "total": len(pending_rows),
                "message": f"Se guardaron {len(pending_rows)} registros correctamente."
            }
        )


    return render(
        request,
        './companies/partials/update_salary_add_masive.html',
        {}
    )


@login_required
@role_required('company', 'accountant')
def update_salary_add_masive_doc(request):
    # Crear el libro de Excel
    wb = Workbook()

    # Hoja principal
    ws = wb.active
    ws.title = "Datos"

    # Columnas fijas
    columns = ["ID Contrato", "Documento", "Nombre", "Nuevo Salario",'Fecha']

    # Escribir encabezados
    for col_num, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_num, value=col_name)
        ws.cell(row=1, column=col_num).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Fijar fila de encabezado
    ws.freeze_panes = "A2"

    # Ajustar ancho de columnas
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(1, len(columns) + 1):
        column_letter = get_column_letter(col)
        dim_holder[column_letter] = ws.column_dimensions[column_letter]
        dim_holder[column_letter].width = 20
    ws.column_dimensions = dim_holder

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ajuste_salarial_masivo.xlsx'

    wb.save(response)
    return response




@login_required
@role_required('company','accountant')
def get_contract_salary(request):
    """
    Endpoint para obtener el salario de un contrato específico
    """
    if request.method == 'GET':
        
        contract_id = request.GET.get('contract_id')
        if not contract_id:
            return JsonResponse({
                'success': False,
                'error': 'ID de contrato requerido'
            })
        
        try:
            contract = Contratos.objects.get(
                idcontrato=contract_id,
                estadocontrato=1  # Solo contratos activos
            )
            return JsonResponse({
                'success': True,
                'salary': contract.salario or 0
            })
        except Contratos.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Contrato no encontrado'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Método no permitido'
    })