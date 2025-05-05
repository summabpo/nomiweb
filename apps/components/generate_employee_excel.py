
import openpyxl
from openpyxl.styles import PatternFill, Alignment
from io import BytesIO

"""
Genera un archivo Excel con los datos de empleados y sus conceptos asociados.

Este script crea un archivo Excel en memoria que contiene un reporte de acumulados de empleados. 
El reporte incluye detalles como el nombre del empleado, los conceptos asociados (id, nombre, cantidad, valor) y los totales por cada empleado.

Parámetros
----------
diccionario_empleados : dict
    Diccionario que contiene la información de los empleados. Cada clave corresponde al identificador de un empleado, y su valor es otro diccionario que contiene detalles como el nombre, los conceptos y el total acumulado.

month1 : str
    Mes inicial para el reporte.

year1 : int
    Año inicial para el reporte.

month2 : str
    Mes final para el reporte.

year2 : int
    Año final para el reporte.

Retorna
-------
bytes
    Un archivo Excel en formato binario (bytes) con los datos de los empleados y sus conceptos asociados.

Descripción
-----------
- El archivo generado tiene un formato con colores diferenciados para los encabezados y los datos de los empleados.
- La primera fila contiene un título "Reporte de acumulados" y la fila siguiente tiene los meses y años de inicio y final del reporte.
- Cada empleado tiene su propio bloque de datos, donde se muestra su nombre, sus conceptos (id, nombre, cantidad, valor) y el total acumulado.
- Al final de cada bloque de empleado, se presenta el total acumulado de ese empleado.
- Las columnas del archivo Excel tienen un ancho adecuado para que los datos sean fácilmente legibles.
- El archivo es guardado en memoria y retornado como un valor en bytes para ser descargado o utilizado en otros contextos.

Notas
-----
- El formato de celdas usa colores personalizados (`azul_fill` y `amarillo_fill`) y alineación centrada para una mejor presentación visual.
- El archivo se guarda con el nombre `excel_temporal.xlsx` en la carpeta `static/docs/` y también se genera en memoria.
"""


output_path = './static/docs/excel_temporal.xlsx'

def generate_employee_excel(diccionario_empleados,month1,year1,month2,year2):
    # Crear un archivo en memoria
    output = BytesIO()
    
    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de acumulados"
    
    # Definir colores y alineación
    azul_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
    amarillo_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    centro_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezados del reporte
    ws["A1"] = "Reporte de acumulados"
    ws.merge_cells("A1:H1")
    
    # Fila de Mes Inicial y Año
    ws["A2"] = "Mes Inicial"
    ws["B2"] = month1
    ws["C2"] = "Año"
    ws["D2"] = year1
    ws["E2"] = "Mes Final"
    ws["F2"] = month2
    ws["G2"] = "Año"
    ws["H2"] = year2
    
    #Estilos para la fila de Mes Inicial y Año
    for cell in ["A2", "E2", "C2", "G2"]:
        ws[cell].fill = azul_fill
    for cell in ["B2", "F2", "D2", "H2"]:
        ws[cell].fill = amarillo_fill
        ws[cell].alignment = centro_alignment
    
    #Datos de empleados
    row_start = 7
    for empleado, info in diccionario_empleados.items():
        # Escribir el nombre del empleado
        ws[f"B{row_start}"] = "Empleado"
        ws[f"C{row_start}"] = info["Empleado"]
        ws.merge_cells(f"C{row_start}:H{row_start}")
        ws[f"C{row_start}"].alignment = centro_alignment
        
        row_start += 1
        
        ws[f"B{row_start}"] = "Id Concepto"
        ws[f"C{row_start}"] = "Nombre Concepto"
        ws[f"D{row_start}"] = "Cantidad"
        ws[f"E{row_start}"] = "Valor"
        ws[f"F{row_start}"] = "Contrato"
        
        ws[f"B{row_start}"].fill = azul_fill
        ws[f"C{row_start}"].fill = azul_fill
        ws[f"D{row_start}"].fill = azul_fill
        ws[f"E{row_start}"].fill = azul_fill
        ws[f"B{row_start}"].alignment = centro_alignment
        ws[f"C{row_start}"].alignment = centro_alignment
        ws[f"D{row_start}"].alignment = centro_alignment
        ws[f"E{row_start}"].alignment = centro_alignment
        
        # Escribir los datos de los conceptos
        for concepto in info["data"]:
            row_start += 1
            ws[f"B{row_start}"] = concepto["idconcepto"]
            ws[f"C{row_start}"] = concepto["nombreconcepto"]
            ws[f"D{row_start}"] = concepto["cantidad"]
            ws[f"E{row_start}"] = concepto["valor"]
            ws[f"F{row_start}"] = info["id"]
            
        row_start += 1
        #Dejar una fila en blanco después de cada empleado
        ws[f"B{row_start}"] = 'Total'
        ws[f"C{row_start}"] = info["total"]
        
        
        row_start += 2
    
    # Ajustar el ancho de las columnas
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    
    wb.save(output_path)
    
    # Guardar el archivo en memoria
    wb.save(output)
    output.seek(0)
    
    
    return output.getvalue()
