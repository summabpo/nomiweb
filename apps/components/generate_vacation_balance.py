import openpyxl
from openpyxl.styles import NamedStyle
from io import BytesIO
from apps.common.models  import Contratos, Vacaciones, Conceptosfijos
from django.db.models import Q, Sum
from django.utils import timezone
from datetime import datetime
from apps.components.utils import calcular_dias_360



def calcular_vacaciones(contrato, concepto, fecha_actual):
    
    """
    Calcula las vacaciones disponibles de un empleado en función de su contrato y los conceptos fijos.

    Parámetros
    ----------
    contrato : Contratos
        El contrato del empleado para el cual se calcularán las vacaciones.
    
    concepto : float
        El porcentaje correspondiente al concepto de vacaciones (por ejemplo, 1.25% para vacaciones).

    fecha_actual : datetime
        La fecha actual (puede ser una fecha proporcionada en el parámetro o la fecha actual del sistema).

    Retorna
    -------
    dict
        Un diccionario con los siguientes valores:
        - 'total_vac_disf' : float : El total de días de vacaciones ya disfrutados.
        - 'total_vac' : float : El total de días de vacaciones que corresponden al empleado según su contrato.
        - 'saldo' : float : El saldo de vacaciones disponible (total vacacional - días disfrutados).
    
    Descripción
    -----------
    Esta función calcula el saldo de vacaciones disponible de un empleado, sumando las vacaciones ya disfrutadas y comparándolas con el total que corresponde según el contrato.
    - Los días disfrutados de vacaciones se obtienen de la tabla `Vacaciones` (tipovac '1' o '2').
    - El total de días de vacaciones se calcula con base en los días del contrato y el porcentaje de vacaciones que corresponde según el concepto (proporcionado como parámetro).
    """

    # Calcular las vacaciones disponibles
    total_vac_disf = Vacaciones.objects.filter(
        idcontrato=contrato.idcontrato
    ).filter(Q(tipovac='1') | Q(tipovac='2')).aggregate(Sum('diasvac'))['diasvac__sum'] or 0

    # Fecha actual y fecha inicial
    fecha_inicial = contrato.fechainiciocontrato

    total_vac = (calcular_dias_360(fecha_inicial.strftime("%Y-%m-%d"), fecha_actual.strftime("%Y-%m-%d"))) * (concepto / 100)
    saldo = total_vac - total_vac_disf

    return {
        "total_vac_disf": round(total_vac_disf, 2),
        "total_vac": round(total_vac, 2),
        "saldo": round(saldo, 2)
    }

def generate_balance_excel(fecha_param=None):
    
    """
    Genera un archivo Excel con el saldo de vacaciones de los empleados para un período determinado.

    Parámetros
    ----------
    fecha_param : str, opcional
        La fecha en formato 'YYYY-MM-DD' para la cual se calculan los saldos de vacaciones. Si no se proporciona, se utiliza la fecha actual.

    Retorna
    -------
    bytes
        Un archivo Excel en formato binario (bytes) con el saldo de vacaciones de cada empleado.

    Descripción
    -----------
    Esta función genera un archivo Excel que contiene un reporte de saldo de vacaciones de los empleados activos, calculando los días disfrutados, los días totales y el saldo de vacaciones disponible.
    
    - Los datos incluidos en el reporte son:
      - Contrato
      - Documento de identidad
      - Nombre completo del empleado
      - Fecha de inicio de contrato
      - Salario
      - Valor de vacaciones (calculado)
      - Vacaciones disfrutadas
      - Total de vacaciones (según contrato)
      - Saldo de vacaciones

    - La fecha actual o la fecha proporcionada se utiliza para calcular el saldo de vacaciones de acuerdo con los días transcurridos desde la fecha de inicio del contrato.
    - Los valores de vacaciones se calculan con base en un concepto fijo almacenado en `Conceptosfijos`.
    - Se aplica formato decimal a las columnas numéricas y formato de fecha a la columna de fecha de inicio de contrato.
    - El archivo es guardado en memoria y retornado como un archivo binario en formato Excel.

    Notas
    -----
    - El cálculo de días de vacaciones es realizado utilizando la función `calcular_dias_360`, que utiliza un cálculo de días en base a 360 días al año.
    - El archivo contiene una hoja llamada "Saldo Vacaciones" con los datos mencionados.
    """

    # Crear un archivo en memoria
    output = BytesIO()

    # Crear un nuevo libro de trabajo y seleccionar la hoja activa
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Saldo Vacaciones"

    # Encabezados del reporte
    headers = ['Contrato', 'Documento', 'Empleado', 'Fecha Contrato', 'Salario', 'Valor', 'Vacaciones Disfrutadas', 'Total Vacaciones', 'Saldo X VAC']
    ws.append(headers)

    # Estilo para formato decimal
    decimal_style = NamedStyle(name='decimal_style', number_format='0.00')

    # Obtener la fecha actual o utilizar la proporcionada
    fecha_actual = timezone.now().date() if fecha_param is None else datetime.strptime(fecha_param, "%Y-%m-%d").date()

    # Obtener el valor fijo del concepto de vacaciones
    concepto = Conceptosfijos.objects.filter(idfijo=9).values_list('valorfijo', flat=True).first()
    valor_fijo = float(concepto)

    # Obtener los contratos activos
    contratos_empleados = Contratos.objects.prefetch_related('idempleado') \
        .filter(estadocontrato=1, tipocontrato__idtipocontrato__in=[1, 2, 3, 4]) \
        .values('idempleado__docidentidad', 'idempleado__sapellido', 'idempleado__papellido',
                'idempleado__pnombre', 'idempleado__snombre', 'idempleado__idempleado',
                'idcontrato', 'fechainiciocontrato', 'salario').order_by('idempleado__papellido')

    # Procesar cada contrato y calcular los valores de vacaciones
    for data in contratos_empleados:
        contrato_id = data['idcontrato']
        contrato = Contratos.objects.get(idcontrato=contrato_id)

        # Calcular las vacaciones
        total_vac_disf = Vacaciones.objects.filter(
            idcontrato=contrato_id
        ).filter(Q(tipovac='1') | Q(tipovac='2')).aggregate(Sum('diasvac'))['diasvac__sum'] or 0

        # Fecha inicial y cálculo de días
        fecha_inicial = contrato.fechainiciocontrato
        total_vac = (calcular_dias_360(fecha_inicial.strftime("%Y-%m-%d"), fecha_actual.strftime("%Y-%m-%d"))) * (valor_fijo / 100)
        saldo = total_vac - total_vac_disf

        # Usar directamente data['fechainiciocontrato'] ya que es de tipo date
        fecha_contrato = data['fechainiciocontrato']

        # Agregar datos a la hoja de trabajo
        ws.append([
            contrato_id,
            data['idempleado__docidentidad'],
            f"{data['idempleado__papellido']} {data['idempleado__sapellido']} {data['idempleado__pnombre']} {data['idempleado__snombre']}",
            fecha_contrato,  # Usar fecha_contrato aquí directamente
            float(data['salario']),
            round(float(data['salario']) / 30, 2) * saldo,
            round(total_vac_disf, 2),
            round(total_vac, 2),
            round(saldo, 2)
        ])

        # Formatear la fecha como string en el formato deseado
        ws.cell(row=ws.max_row, column=4).number_format = 'DD/MM/YYYY'

    # Aplicar formato decimal a las columnas relevantes
    for col in ['E', 'F', 'G', 'H', 'I']:
        for cell in ws[col]:
            cell.style = decimal_style

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Guardar el archivo en memoria
    wb.save(output)
    output.seek(0)

    return output.getvalue()
