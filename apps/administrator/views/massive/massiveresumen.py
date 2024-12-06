from django.shortcuts import render,redirect
from django.contrib import messages
from apps.common.models import Tipodocumento ,Ciudades , Paises ,Empresa , Contratosemp
import openpyxl
from datetime import datetime
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import default_storage
# Create your views here.

def massiveresumen(request):
    visual = False
    migration_status = []  # Lista para almacenar mensajes de éxito o error
    migration_count = 0  # Contador de migraciones exitosas
    error_count = 0    # Contador de errores
    reporte = []
    errors = []
    if request.method == "POST":
        visual = True
        excel_file = request.FILES.get('excel_file')
        
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
        except Exception as e:
            messages.success(request, f"Error al procesar el archivo: {e}")
            return redirect('admin:massiveresumen')
        
            
            
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                
                migration_count += 1
                # Extrae los datos de cada columna del archivo Excel
                docidentidad = row[0]  # Columna 0: docidentidad
                pnombre = str(row[1])  # Columna 1: pnombre
                snombre = str(row[2])  # Columna 2: snombre
                papellido = str(row[3])  # Columna 3: papellido
                sapellido = str(row[4])  # Columna 4: sapellido
                email = str(row[5])  # Columna 5: email
                telefonoempleado = str(row[6]).strip()  # Columna 6: telefonoempleado
                celular = str(row[7])  # Columna 7: celular
                direccionempleado = str(row[8])  # Columna 8: direccionempleado
                sexo = row[9]  # Columna 9: sexo
                fechanac = row[10]  # Columna 10: fechanac
                estadocivil = row[11]  # Columna 11: estadocivil
                profesion = row[12]  # Columna 12: profesion
                niveleducativo = row[13]  # Columna 13: niveleducativo
                estatura = row[14]  # Columna 14: estatura
                peso = row[15]  # Columna 15: peso
                gruposanguineo = row[16]  # Columna 16: gruposanguineo
                fechaexpedicion = row[17]  # Columna 17: fechaexpedicion
                formatohv = row[18]  # Columna 18: formatohv
                dotpantalon = row[19]  # Columna 19: dotpantalon
                dotcamisa = row[20]  # Columna 20: dotcamisa
                dotzapatos = row[21]  # Columna 21: dotzapatos
                estadocontrato = row[22]  # Columna 22: estadocontrato
                estrato = row[23]  # Columna 23: estrato
                numlibretamil = row[24]  # Columna 24: numlibretamil
                fotografiaempleado = row[25]  # Columna 25: fotografiaempleado
                ciudadexpedicion_id = row[26]  # Columna 26: ciudadexpedicion_id
                ciudadnacimiento_id = row[27]  # Columna 27: ciudadnacimiento_id
                ciudadresidencia_id = row[28]  # Columna 28: ciudadresidencia_id
                id_empresa_id = row[29]  # Columna 29: id_empresa_id
                paisnacimiento_id = row[30]  # Columna 30: paisnacimiento_id
                paisresidencia_id = row[31]  # Columna 31: paisresidencia_id
                tipodocident_id = row[32]  # Columna 32: tipodocident_id
                        
                # Validación de cada campo según el tipo de dato

                if not isinstance(docidentidad, int):
                    # Aquí puedes generar el reporte para 'docidentidad'
                    reporte.append(f"El valor de 'docidentidad' en la fila {i} no es un número entero.")
                    error_count += 1

                if not isinstance(pnombre, str):
                    reporte.append(f"El valor de 'pnombre' en la fila {i} no es un texto válido (no es un string).")
                    error_count += 1
                elif len(pnombre) > 50:
                    reporte.append(f"El valor de 'pnombre' en la fila {i} excede los 50 caracteres.")
                    error_count += 1

                if snombre and (not isinstance(snombre, str) or len(snombre) > 50):
                    # Validación de 'snombre', si existe, debe ser un string y máximo 50 caracteres
                    reporte.append(f"El valor de 'snombre' en la fila {i} no es un texto válido o excede los 50 caracteres.")
                    error_count += 1

                if not isinstance(papellido, str) or len(papellido) > 50:
                    # Validación de 'papellido'
                    reporte.append(f"El valor de 'papellido' en la fila {i} no es un texto válido o excede los 50 caracteres.")
                    error_count += 1

                if sapellido and (not isinstance(sapellido, str) or len(sapellido) > 50):
                    # Validación de 'sapellido'
                    reporte.append(f"El valor de 'sapellido' en la fila {i} no es un texto válido o excede los 50 caracteres.")
                    error_count += 1

                if email and (not isinstance(email, str) or len(email) > 255):
                    #Validación de 'email'
                    reporte.append(f"El valor de 'email' en la fila {i} no es un texto válido o excede los 255 caracteres.")
                    error_count += 1

                if not isinstance(telefonoempleado, str):
                    reporte.append(f"El valor de 'telefonoempleado' en la fila {i} no es un texto válido (no es un string).")
                    error_count += 1
                elif len(telefonoempleado) > 12:
                    reporte.append(f"El valor de 'telefonoempleado' en la fila {i} excede los 12 caracteres.")
                    error_count += 1
                
                if celular and (not isinstance(celular, str) or len(celular) > 12):
                    #Validación de 'celular'
                    reporte.append(f"El valor de 'celular' en la fila {i} no es un texto válido o excede los 12 caracteres.")
                    error_count += 1

                if direccionempleado and (not isinstance(direccionempleado, str) or len(direccionempleado) > 100):
                    #Validación de 'direccionempleado'
                    reporte.append(f"El valor de 'direccionempleado' en la fila {i} no es un texto válido o excede los 100 caracteres.")
                    error_count += 1

                if sexo and (not isinstance(sexo, str) or len(sexo) > 255):
                    #Validación de 'sexo'
                    reporte.append(f"El valor de 'sexo' en la fila {i} no es un texto válido o excede los 255 caracteres.")
                    error_count += 1

                try:
                    fechanac = datetime.strptime(fechanac, "%Y-%m-%d").date()
                except ValueError:
                    reporte.append(f"El valor de 'fechanac' en la fila {i} no es una fecha válida.")
                    error_count += 1

                #Validación de las claves foráneas, si son números enteros válidos
                if not isinstance(ciudadnacimiento_id, int) or not Ciudades.objects.filter(idciudad=ciudadnacimiento_id).exists():
                    reporte.append(f"El valor de 'ciudadnacimiento_id' en la fila {i} no es válido o no existe.")
                    error_count += 1

                if not isinstance(paisnacimiento_id, int) or not Paises.objects.filter(idpais=paisnacimiento_id).exists():
                    reporte.append(f"El valor de 'paisnacimiento_id' en la fila {i} no es válido o no existe.")
                    error_count += 1

                if not isinstance(ciudadresidencia_id, int) or not Ciudades.objects.filter(idciudad=ciudadresidencia_id).exists():
                    reporte.append(f"El valor de 'ciudadresidencia_id' en la fila {i} no es válido o no existe.")
                    error_count += 1

                if not isinstance(paisresidencia_id, int) or not Paises.objects.filter(idpais=paisresidencia_id).exists():
                    reporte.append(f"El valor de 'paisresidencia_id' en la fila {i} no es válido o no existe.")
                    error_count += 1

                if estadocivil and (not isinstance(estadocivil, str) or len(estadocivil) > 20):
                    reporte.append(f"El valor de 'estadocivil' en la fila {i} no es válido o excede los 20 caracteres.")
                    error_count += 1

                if profesion and (not isinstance(profesion, str) or len(profesion) > 180):
                    reporte.append(f"El valor de 'profesion' en la fila {i} no es válido o excede los 180 caracteres.")
                    error_count += 1

                if niveleducativo and (not isinstance(niveleducativo, str) or len(niveleducativo) > 25):
                    reporte.append(f"El valor de 'niveleducativo' en la fila {i} no es válido o excede los 25 caracteres.")
                    error_count += 1

                if estatura and (not isinstance(estatura, str) or len(estatura) > 10):
                    reporte.append(f"El valor de 'estatura' en la fila {i} no es válido o excede los 10 caracteres.")
                    error_count += 1

                if peso and (not isinstance(peso, str) or len(peso) > 10):
                    reporte.append(f"El valor de 'peso' en la fila {i} no es válido o excede los 10 caracteres.")
                    error_count += 1

                if gruposanguineo and (not isinstance(gruposanguineo, str) or len(gruposanguineo) > 10):
                    reporte.append(f"El valor de 'gruposanguineo' en la fila {i} no es válido o excede los 10 caracteres.")
                    error_count += 1

                try:
                    fechaexpedicion = datetime.strptime(fechaexpedicion, "%Y-%m-%d").date()
                except ValueError:
                    reporte.append(f"El valor de 'fechaexpedicion' en la fila {i} no es una fecha válida.")
                    error_count += 1

                if not isinstance(ciudadexpedicion_id, int) or not Ciudades.objects.filter(idciudad=ciudadexpedicion_id).exists():
                    reporte.append(f"El valor de 'ciudadexpedicion_id' en la fila {i} no es válido o no existe.")
                    error_count += 1

                if formatohv and (not isinstance(formatohv, str) or len(formatohv) > 25):
                    reporte.append(f"El valor de 'formatohv' en la fila {i} no es válido o excede los 25 caracteres.")
                    error_count += 1

                if dotpantalon and (not isinstance(dotpantalon, str) or len(dotpantalon) > 10):
                    reporte.append(f"El valor de 'dotpantalon' en la fila {i} no es válido o excede los 10 caracteres.")
                    error_count += 1

                if dotcamisa and (not isinstance(dotcamisa, str) or len(dotcamisa) > 10):
                    reporte.append(f"El valor de 'dotcamisa' en la fila {i} no es válido o excede los 10 caracteres.")
                    error_count += 1

                if dotzapatos and (not isinstance(dotzapatos, str) or len(dotzapatos) > 10):
                    reporte.append(f"El valor de 'dotzapatos' en la fila {i} no es válido o excede los 10 caracteres.")
                    error_count += 1

                if estadocontrato and (not isinstance(estadocontrato, int) or estadocontrato < 0):
                    reporte.append(f"El valor de 'estadocontrato' en la fila {i} no es válido.")
                    error_count += 1

                if estrato and (not isinstance(estrato, str) or len(estrato) > 5):
                    reporte.append(f"El valor de 'estrato' en la fila {i} no es válido o excede los 5 caracteres.")
                    error_count += 1

                if numlibretamil and (not isinstance(numlibretamil, str) or len(numlibretamil) > 10):
                    reporte.append(f"El valor de 'numlibretamil' en la fila {i} no es válido o excede los 10 caracteres.")
                    error_count += 1

                if fotografiaempleado and (not isinstance(fotografiaempleado, str) or len(fotografiaempleado) > 25):
                    reporte.append(f"El valor de 'fotografiaempleado' en la fila {i} no es válido o excede los 25 caracteres.")
                    error_count += 1

                if not isinstance(id_empresa_id, int) or not Empresa.objects.filter(idempresa=id_empresa_id).exists():
                    reporte.append(f"El valor de 'id_empresa_id' en la fila {i} no es válido o no existe.")
                    error_count += 1
                
                if not isinstance(ciudadnacimiento_id, int) or not Ciudades.objects.filter(idciudad=ciudadnacimiento_id).exists():
                    reporte.append(f"El valor de 'ciudadnacimiento_id' en la fila {i} no es válido o no existe.")
                    error_count += 1
                
            except Exception as e:
                errors.append(f"Fila {i}: Error de procesamiento - {str(e)}")
                
    context = {
        'visual' :visual ,
        'reporte': reporte,
        'error_count': error_count,
        'total' : migration_count
        
    }
    return render(request, './admin/massiveresumen.html',context )



@csrf_exempt
def massiveresumenmigrate(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        archivo = request.FILES['excel_file']
        
        if archivo:
            try:
                workbook = openpyxl.load_workbook(archivo)
                sheet = workbook.active
            except Exception as e:
                return JsonResponse({'success': False, 'message': 'Error al procesar el archivo: {e}'})
        
            # Ejemplo: procesar el archivo
            try:
                # Suponiendo que el archivo sea un Excel y estés usando pandas
                reporte = []
                error_count = 0
                migration_count = 0

                # Iteramos sobre cada fila del DataFrame de pandas
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        migration_count += 1
                        docidentidad = row[0]  # Columna 0: docidentidad
                        pnombre = str(row[1])  # Columna 1: pnombre
                        snombre = str(row[2])  # Columna 2: snombre
                        papellido = str(row[3])  # Columna 3: papellido
                        sapellido = str(row[4])  # Columna 4: sapellido
                        email = str(row[5])  # Columna 5: email
                        telefonoempleado = str(row[6]).strip()  # Columna 6: telefonoempleado
                        celular = str(row[7])  # Columna 7: celular
                        direccionempleado = str(row[8])  # Columna 8: direccionempleado
                        sexo = row[9]  # Columna 9: sexo
                        fechanac = row[10]  # Columna 10: fechanac
                        estadocivil = row[11]  # Columna 11: estadocivil
                        profesion = row[12]  # Columna 12: profesion
                        niveleducativo = row[13]  # Columna 13: niveleducativo
                        estatura = row[14]  # Columna 14: estatura
                        peso = row[15]  # Columna 15: peso
                        gruposanguineo = row[16]  # Columna 16: gruposanguineo
                        fechaexpedicion = row[17]  # Columna 17: fechaexpedicion
                        formatohv = row[18]  # Columna 18: formatohv
                        dotpantalon = row[19]  # Columna 19: dotpantalon
                        dotcamisa = row[20]  # Columna 20: dotcamisa
                        dotzapatos = row[21]  # Columna 21: dotzapatos
                        estadocontrato = row[22]  # Columna 22: estadocontrato
                        estrato = row[23]  # Columna 23: estrato
                        numlibretamil = row[24]  # Columna 24: numlibretamil
                        fotografiaempleado = row[25]  # Columna 25: fotografiaempleado
                        ciudadexpedicion_id = row[26]  # Columna 26: ciudadexpedicion_id
                        ciudadnacimiento_id = row[27]  # Columna 27: ciudadnacimiento_id
                        ciudadresidencia_id = row[28]  # Columna 28: ciudadresidencia_id
                        id_empresa_id = row[29]  # Columna 29: id_empresa_id
                        paisnacimiento_id = row[30]  # Columna 30: paisnacimiento_id
                        paisresidencia_id = row[31]  # Columna 31: paisresidencia_id
                        tipodocident_id = row[32]  # Columna 32: tipodocident_id
                        
                        fecha1 = fechanac  
                        fecha2 = fechaexpedicion 
                        
                            
                        # Validación de cada campo según el tipo de dato

                        if not isinstance(docidentidad, int):
                            # Aquí puedes generar el reporte para 'docidentidad'
                            reporte.append(f"El valor de 'docidentidad' en la fila {i} no es un número entero.")
                            error_count += 1

                        if not isinstance(pnombre, str):
                            reporte.append(f"El valor de 'pnombre' en la fila {i} no es un texto válido (no es un string).")
                            error_count += 1
                        elif len(pnombre) > 50:
                            reporte.append(f"El valor de 'pnombre' en la fila {i} excede los 50 caracteres.")
                            error_count += 1

                        if snombre and (not isinstance(snombre, str) or len(snombre) > 50):
                            # Validación de 'snombre', si existe, debe ser un string y máximo 50 caracteres
                            reporte.append(f"El valor de 'snombre' en la fila {i} no es un texto válido o excede los 50 caracteres.")
                            error_count += 1

                        if not isinstance(papellido, str) or len(papellido) > 50:
                            # Validación de 'papellido'
                            reporte.append(f"El valor de 'papellido' en la fila {i} no es un texto válido o excede los 50 caracteres.")
                            error_count += 1

                        if sapellido and (not isinstance(sapellido, str) or len(sapellido) > 50):
                            # Validación de 'sapellido'
                            reporte.append(f"El valor de 'sapellido' en la fila {i} no es un texto válido o excede los 50 caracteres.")
                            error_count += 1

                        if email and (not isinstance(email, str) or len(email) > 255):
                            #Validación de 'email'
                            reporte.append(f"El valor de 'email' en la fila {i} no es un texto válido o excede los 255 caracteres.")
                            error_count += 1

                        if not isinstance(telefonoempleado, str):
                            reporte.append(f"El valor de 'telefonoempleado' en la fila {i} no es un texto válido (no es un string).")
                            error_count += 1
                        elif len(telefonoempleado) > 12:
                            reporte.append(f"El valor de 'telefonoempleado' en la fila {i} excede los 12 caracteres.")
                            error_count += 1
                        
                        if celular and (not isinstance(celular, str) or len(celular) > 12):
                            #Validación de 'celular'
                            reporte.append(f"El valor de 'celular' en la fila {i} no es un texto válido o excede los 12 caracteres.")
                            error_count += 1

                        if direccionempleado and (not isinstance(direccionempleado, str) or len(direccionempleado) > 100):
                            #Validación de 'direccionempleado'
                            reporte.append(f"El valor de 'direccionempleado' en la fila {i} no es un texto válido o excede los 100 caracteres.")
                            error_count += 1

                        if sexo and (not isinstance(sexo, str) or len(sexo) > 255):
                            #Validación de 'sexo'
                            reporte.append(f"El valor de 'sexo' en la fila {i} no es un texto válido o excede los 255 caracteres.")
                            error_count += 1


                        #Validación de las claves foráneas, si son números enteros válidos
                        if not isinstance(ciudadnacimiento_id, int) or not Ciudades.objects.filter(idciudad=ciudadnacimiento_id).exists():
                            reporte.append(f"El valor de 'ciudadnacimiento_id' en la fila {i} no es válido o no existe.")
                            error_count += 1

                        if not isinstance(paisnacimiento_id, int) or not Paises.objects.filter(idpais=paisnacimiento_id).exists():
                            reporte.append(f"El valor de 'paisnacimiento_id' en la fila {i} no es válido o no existe.")
                            error_count += 1

                        if not isinstance(ciudadresidencia_id, int) or not Ciudades.objects.filter(idciudad=ciudadresidencia_id).exists():
                            reporte.append(f"El valor de 'ciudadresidencia_id' en la fila {i} no es válido o no existe.")
                            error_count += 1

                        if not isinstance(paisresidencia_id, int) or not Paises.objects.filter(idpais=paisresidencia_id).exists():
                            reporte.append(f"El valor de 'paisresidencia_id' en la fila {i} no es válido o no existe.")
                            error_count += 1

                        if estadocivil and (not isinstance(estadocivil, str) or len(estadocivil) > 20):
                            reporte.append(f"El valor de 'estadocivil' en la fila {i} no es válido o excede los 20 caracteres.")
                            error_count += 1

                        if profesion and (not isinstance(profesion, str) or len(profesion) > 180):
                            reporte.append(f"El valor de 'profesion' en la fila {i} no es válido o excede los 180 caracteres.")
                            error_count += 1

                        if niveleducativo and (not isinstance(niveleducativo, str) or len(niveleducativo) > 25):
                            reporte.append(f"El valor de 'niveleducativo' en la fila {i} no es válido o excede los 25 caracteres.")
                            error_count += 1

                        if estatura and (not isinstance(estatura, str) or len(estatura) > 10):
                            reporte.append(f"El valor de 'estatura' en la fila {i} no es válido o excede los 10 caracteres.")
                            error_count += 1

                        if peso and (not isinstance(peso, str) or len(peso) > 10):
                            reporte.append(f"El valor de 'peso' en la fila {i} no es válido o excede los 10 caracteres.")
                            error_count += 1

                        if gruposanguineo and (not isinstance(gruposanguineo, str) or len(gruposanguineo) > 10):
                            reporte.append(f"El valor de 'gruposanguineo' en la fila {i} no es válido o excede los 10 caracteres.")
                            error_count += 1


                        if not isinstance(ciudadexpedicion_id, int) or not Ciudades.objects.filter(idciudad=ciudadexpedicion_id).exists():
                            reporte.append(f"El valor de 'ciudadexpedicion_id' en la fila {i} no es válido o no existe.")
                            error_count += 1

                        if formatohv and (not isinstance(formatohv, str) or len(formatohv) > 25):
                            reporte.append(f"El valor de 'formatohv' en la fila {i} no es válido o excede los 25 caracteres.")
                            error_count += 1

                        if dotpantalon and (not isinstance(dotpantalon, str) or len(dotpantalon) > 10):
                            reporte.append(f"El valor de 'dotpantalon' en la fila {i} no es válido o excede los 10 caracteres.")
                            error_count += 1

                        if dotcamisa and (not isinstance(dotcamisa, str) or len(dotcamisa) > 10):
                            reporte.append(f"El valor de 'dotcamisa' en la fila {i} no es válido o excede los 10 caracteres.")
                            error_count += 1

                        if dotzapatos and (not isinstance(dotzapatos, str) or len(dotzapatos) > 10):
                            reporte.append(f"El valor de 'dotzapatos' en la fila {i} no es válido o excede los 10 caracteres.")
                            error_count += 1

                        if estadocontrato and (not isinstance(estadocontrato, int) or estadocontrato < 0):
                            reporte.append(f"El valor de 'estadocontrato' en la fila {i} no es válido.")
                            error_count += 1

                        if estrato and (not isinstance(estrato, str) or len(estrato) > 5):
                            reporte.append(f"El valor de 'estrato' en la fila {i} no es válido o excede los 5 caracteres.")
                            error_count += 1

                        if numlibretamil and (not isinstance(numlibretamil, str) or len(numlibretamil) > 10):
                            reporte.append(f"El valor de 'numlibretamil' en la fila {i} no es válido o excede los 10 caracteres.")
                            error_count += 1

                        if fotografiaempleado and (not isinstance(fotografiaempleado, str) or len(fotografiaempleado) > 25):
                            reporte.append(f"El valor de 'fotografiaempleado' en la fila {i} no es válido o excede los 25 caracteres.")
                            error_count += 1

                        if not isinstance(id_empresa_id, int) or not Empresa.objects.filter(idempresa=id_empresa_id).exists():
                            reporte.append(f"El valor de 'id_empresa_id' en la fila {i} no es válido o no existe.")
                            error_count += 1
                            
                        if not isinstance(tipodocident_id, int) or not Tipodocumento.objects.filter(id_tipo_doc=tipodocident_id).exists():
                            reporte.append(f"El valor de 'tipodocident_id' en la fila {i} no es válido o no existe.")
                            error_count += 1

                        
                        # Si todos los campos son válidos, creamos el contrato
                        if error_count == 0:
                            typedoc = Tipodocumento.objects.get(id_tipo_doc = tipodocident_id )
                            cnaci = Ciudades.objects.get(idciudad = ciudadnacimiento_id )
                            cresi = Ciudades.objects.get(idciudad = ciudadresidencia_id )
                            cexpe = Ciudades.objects.get(idciudad = ciudadexpedicion_id )
                            
                            pnaci = Paises.objects.get(idpais = paisnacimiento_id )
                            presi = Paises.objects.get(idpais = paisresidencia_id ) 
                            
                            empr = Empresa.objects.get( idempresa = id_empresa_id ) 
                            
                            
                            
                                                        
                            contrato = Contratosemp.objects.create(
                                docidentidad=docidentidad,
                                tipodocident = typedoc,
                                pnombre=pnombre,
                                snombre=snombre,
                                papellido=papellido,
                                sapellido=sapellido,
                                email=email,
                                telefonoempleado=telefonoempleado,
                                celular=celular,
                                direccionempleado=direccionempleado,
                                sexo=sexo,
                                fechanac=fecha1,
                                ciudadnacimiento = cnaci ,
                                paisnacimiento = pnaci,
                                ciudadresidencia = cresi,
                                paisresidencia = presi,
                                estadocivil=estadocivil,
                                profesion=profesion,
                                niveleducativo=niveleducativo,
                                estatura=estatura,
                                peso=peso,
                                gruposanguineo=gruposanguineo,
                                fechaexpedicion=fecha2,
                                ciudadexpedicion = cexpe ,
                                formatohv=formatohv,
                                dotpantalon=dotpantalon,
                                dotcamisa=dotcamisa,
                                dotzapatos=dotzapatos,
                                estadocontrato=estadocontrato,
                                estrato=estrato,
                                numlibretamil=numlibretamil,
                                fotografiaempleado=fotografiaempleado,
                                id_empresa = empr ,
                            )
                            
                    except Exception as e:
                        reporte.append(f"Error al procesar la fila {i}: {str(e)}")
                        print(str(e))
                        error_count += 1
                
                if error_count > 0:
                    print('llege aqui 1 ')
                    return JsonResponse({'success': False, 'message': f'Error en la validación de {error_count} filas.', 'reporte': reporte})
                    
                else:
                    print('llege aqui 2 ')
                    return JsonResponse({'success': True, 'message': f'{migration_count} contratos migrados con éxito.'})
            
            except Exception as e:
                print('llege aqui 3 ')
                return JsonResponse({'success': False, 'message': str(e)})

        else:
            print('llege aqui 4 ')
            return JsonResponse({'success': False, 'message': 'No se ha recibido un archivo válido'})
    else:
        print('llege aqui 5 ')
        return JsonResponse({'success': False, 'message': 'Método inválido o archivo no encontrado'})




